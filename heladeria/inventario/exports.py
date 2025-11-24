# inventario/exports.py
from io import BytesIO
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet

# =====================================================================
#  UTILIDAD BASE
# =====================================================================
def _excel_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    resp = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f"attachment; filename={filename}"
    return resp

def _pdf_response(buffer, filename):
    buffer.seek(0)
    resp = HttpResponse(buffer, content_type="application/pdf")
    resp["Content-Disposition"] = f"attachment; filename={filename}"
    return resp


# =====================================================================
#  EXPORTAR DISPONIBILIDAD (ya tenías esto)
# =====================================================================

def exportar_disponibilidad_csv(qs, ctx):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=reporte_disponibilidad.csv"
    writer = csv.writer(response)

    writer.writerow([
        "Categoría",
        "Insumo",
        "Unidad",
        "Precio Unitario",
        "Stock Total",
        "Lotes con Stock",
        "Próx. Vencimiento",
    ])

    for cat in ctx["categorias"]:
        for ins in cat["insumos"]:
            writer.writerow([
                cat["categoria"].nombre,
                ins.nombre,
                ins.unidad_medida.nombre if ins.unidad_medida else "",
                float(ins.precio_unitario or 0),
                float(ins.stock_total or 0),
                len(getattr(ins, "lotes_vis", []) or []),
                ins.prox_vencimiento.strftime("%Y-%m-%d") if ins.prox_vencimiento else "—",
            ])

    # totales igual que PDF
    writer.writerow([])
    writer.writerow(["Stock total:", f"{ctx['total_stock']:.2f}"])
    writer.writerow(["Precio total:", f"{ctx['total_valor']:.0f}"])

    return response


def exportar_disponibilidad_excel(qs, ctx):
    wb = Workbook()
    ws = wb.active
    ws.title = "Disponibilidad"

    headers = [
        "Categoría",
        "Insumo",
        "Unidad",
        "Precio Unitario",
        "Stock Total",
        "Lotes con Stock",
        "Próx. Vencimiento",
    ]
    ws.append(headers)

    for cat in ctx["categorias"]:
        for ins in cat["insumos"]:
            ws.append([
                cat["categoria"].nombre,
                ins.nombre,
                ins.unidad_medida.nombre if ins.unidad_medida else "",
                float(ins.precio_unitario or 0),
                float(ins.stock_total or 0),
                len(getattr(ins, "lotes_vis", []) or []),
                ins.prox_vencimiento.strftime("%Y-%m-%d") if ins.prox_vencimiento else "—",
            ])

    # línea en blanco y totales igual que PDF
    ws.append([])
    ws.append(["Stock total:", float(ctx["total_stock"])])
    ws.append(["Precio total:", float(ctx["total_valor"])])

    return _excel_response(wb, "reporte_disponibilidad.xlsx")


def exportar_disponibilidad_pdf(qs, ctx):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        title="Reporte Disponibilidad de Insumos"
    )

    styles = getSampleStyleSheet()
    elements = []

    # Título igual al PDF objetivo
    elements.append(
        Paragraph(
            f"<b>Reporte de Disponibilidad de Insumos - {ctx['today']:%Y-%m-%d}</b>",
            styles["Title"]
        )
    )
    elements.append(Paragraph(" ", styles["Normal"]))

    # Por cada categoría, mismo formato que el PDF
    for bloque in ctx["categorias"]:
        cat = bloque["categoria"].nombre

        elements.append(Paragraph(f"<b>Categoría: {cat}</b>", styles["Heading3"]))

        data = [[
            "Insumo",
            "Unidad",
            "Precio Unitario",
            "Stock Total",
            "Lotes con Stock",
            "Próx. Vencimiento",
        ]]

        for ins in bloque["insumos"]:
            lotes_count = len(getattr(ins, "lotes_vis", []) or [])
            prox = ins.prox_vencimiento.strftime("%Y-%m-%d") if ins.prox_vencimiento else "—"

            data.append([
                ins.nombre,
                ins.unidad_medida.nombre if ins.unidad_medida else "",
                f"{ins.precio_unitario:,.0f}".replace(",", "."),
                f"{(ins.stock_total or 0):.2f}",
                str(lotes_count),
                prox,
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0b3861")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("ALIGN", (2,1), (-2,-1), "RIGHT"),
            ("ALIGN", (-1,1), (-1,-1), "CENTER"),
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.lightgrey]),
        ]))

        elements.append(table)
        elements.append(Paragraph(" ", styles["Normal"]))

    # Totales igual al PDF objetivo
    elements.append(Paragraph(f"<b>Stock total: {ctx['total_stock']:.2f}</b>", styles["Normal"]))
    elements.append(Paragraph(f"<b>Precio total: {ctx['total_valor']:,.0f}</b>".replace(",", "."), styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=reporte_disponibilidad.pdf"
    return response


# =====================================================================
#  EXPORTAR PRÓXIMOS A VENCER
# =====================================================================
def exportar_proximos_excel(qs, dias):
    wb = Workbook()
    ws = wb.active
    ws.title = "Próximos a vencer"

    ws.append(["Insumo", "Bodega", "F. Ingreso", "F. Expiración", "Cant. Inicial", "Cant. Actual"])

    for l in qs:
        ws.append([
            l.insumo.nombre,
            l.bodega.nombre,
            str(l.fecha_ingreso),
            str(l.fecha_expiracion),
            float(l.cantidad_inicial or 0),
            float(l.cantidad_actual or 0)
        ])

    return _excel_response(wb, f"lotes_proximos_{dias}_dias.xlsx")


def exportar_proximos_pdf(qs, dias):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)

    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph(f"<b>Próximos a Vencer ({dias} días)</b>", styles["Title"]))

    data = [["Insumo", "Bodega", "F. Ing.", "F. Exp.", "Inicial", "Actual"]]
    for l in qs:
        data.append([
            l.insumo.nombre,
            l.bodega.nombre,
            str(l.fecha_ingreso),
            str(l.fecha_expiracion),
            str(l.cantidad_inicial),
            str(l.cantidad_actual)
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#006699")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
    ]))

    elements.append(table)
    doc.build(elements)
    return _pdf_response(buffer, f"lotes_proximos_{dias}_dias.pdf")


# =====================================================================
#  EXPORTAR MOVIMIENTOS
# =====================================================================
def exportar_mov_excel(entradas, salidas, q):
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    ws.append(["Tipo", "Fecha", "Insumo", "Ubicación", "Cantidad", "Observaciones"])

    for e in entradas:
        ws.append(["Entrada", str(e.fecha), e.insumo.nombre, e.ubicacion.nombre,
                   float(e.cantidad), e.observaciones])

    for s in salidas:
        ws.append(["Salida", str(s.fecha_generada), s.insumo.nombre, s.ubicacion.nombre,
                   float(s.cantidad), s.observaciones])

    return _excel_response(wb, "movimientos.xlsx")


def exportar_mov_pdf(entradas, salidas, q):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>Historial de Movimientos</b>", styles["Title"]))

    data = [["Tipo", "Fecha", "Insumo", "Ubicación", "Cantidad", "Obs."]]

    for e in entradas:
        data.append(["Entrada", str(e.fecha), e.insumo.nombre, e.ubicacion.nombre,
                     str(e.cantidad), e.observaciones])

    for s in salidas:
        data.append(["Salida", str(s.fecha_generada), s.insumo.nombre, s.ubicacion.nombre,
                     str(s.cantidad), s.observaciones])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.darkblue),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)

    return _pdf_response(buffer, "movimientos.pdf")


# =====================================================================
#  EXPORTAR STOCK CONSOLIDADO
# =====================================================================
def exportar_stock_excel(qs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Consolidado"

    ws.append(["Insumo", "Categoría", "Stock Total", "Mínimo", "Máximo"])

    for i in qs:
        ws.append([
            i.nombre,
            i.categoria.nombre,
            float(i.stock_total),
            float(i.stock_min_storage),
            float(i.stock_max_storage)
        ])

    return _excel_response(wb, "stock_consolidado.xlsx")


def exportar_stock_pdf(qs):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph("<b>Stock Consolidado</b>", styles["Title"]))

    data = [["Insumo", "Categoría", "Total", "Mínimo", "Máximo"]]

    for i in qs:
        data.append([
            i.nombre,
            i.categoria.nombre,
            str(i.stock_total),
            str(i.stock_min_storage),
            str(i.stock_max_storage),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.green),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)

    return _pdf_response(buffer, "stock_consolidado.pdf")


# =====================================================================
#  EXPORTAR ÓRDENES
# =====================================================================
def exportar_ordenes_excel(qs, estado):
    wb = Workbook()
    ws = wb.active
    ws.title = "Órdenes"

    ws.append(["ID", "Fecha", "Solicitante", "Estado", "Total Ítems"])

    for o in qs:
        ws.append([
            o.id,
            str(o.fecha),
            o.usuario.username,
            o.estado,
            o.detalles.count()
        ])

    return _excel_response(wb, f"ordenes_{estado or 'todas'}.xlsx")


def exportar_ordenes_pdf(qs, estado):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>Órdenes</b>", styles["Title"]))

    data = [["ID", "Fecha", "Solicitante", "Estado", "Items"]]

    for o in qs:
        data.append([
            str(o.id),
            str(o.fecha),
            o.usuario.username,
            o.estado,
            str(o.detalles.count())
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.brown),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)

    return _pdf_response(buffer, f"ordenes_{estado or 'todas'}.pdf")


# =====================================================================
#  EXPORTAR LOTES (para /inventario/lotes/exportar/)
# =====================================================================
def exportar_lotes_excel(qs, meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lotes"

    ws.append([
        "Insumo", "Bodega", "Proveedor",
        "F. Ingreso", "F. Expiración",
        "Cant. Inicial", "Cant. Actual"
    ])

    for l in qs:
        ws.append([
            l.insumo.nombre if l.insumo else "",
            l.bodega.nombre if l.bodega else "",
            getattr(l.proveedor, "nombre", "") if hasattr(l, "proveedor") else "",
            str(l.fecha_ingreso or ""),
            str(l.fecha_expiracion or ""),
            float(getattr(l, "cant_ini", l.cantidad_inicial or 0)),
            float(getattr(l, "cant_act", l.cantidad_actual or 0)),
        ])

    return _excel_response(wb, "lotes_insumo.xlsx")


def exportar_lotes_pdf(qs, meta):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    titulo = "Lotes de Insumo"
    if meta.get("filtro_vencimiento") == "proximos":
        titulo += f" (Próximos a vencer {meta.get('dias_proximos', 30)} días)"
    elements.append(Paragraph(f"<b>{titulo}</b>", styles["Title"]))

    data = [[
        "Insumo", "Bodega", "Proveedor",
        "F. Ing.", "F. Exp.",
        "Inicial", "Actual"
    ]]

    for l in qs:
        data.append([
            l.insumo.nombre if l.insumo else "",
            l.bodega.nombre if l.bodega else "",
            getattr(l.proveedor, "nombre", "") if hasattr(l, "proveedor") else "",
            str(l.fecha_ingreso or ""),
            str(l.fecha_expiracion or ""),
            str(getattr(l, "cant_ini", l.cantidad_inicial or 0)),
            str(getattr(l, "cant_act", l.cantidad_actual or 0)),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0b3861")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
    ]))

    elements.append(table)
    doc.build(elements)
    return _pdf_response(buffer, "lotes_insumo.pdf")