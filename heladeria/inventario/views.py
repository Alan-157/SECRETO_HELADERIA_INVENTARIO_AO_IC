from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from accounts.decorators import perfil_required
from accounts.services import user_has_role
from django.http import JsonResponse, HttpResponseBadRequest
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Q, DecimalField, Count, Min, F, Prefetch
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.forms import ModelForm, inlineformset_factory, formset_factory
from django.views.decorators.http import require_POST
from accounts.services import user_has_role
from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch
from .services import check_and_create_stock_alerts
from .models import (
    Insumo, Categoria, Bodega,
    Entrada, Salida, InsumoLote,
    OrdenInsumo, OrdenInsumoDetalle, UnidadMedida, AlertaInsumo, Proveedor,
    ESTADO_ORDEN_CHOICES, TIPO_ORDEN_CHOICES            
)
from django.urls import reverse
from .forms import (
    InsumoForm, CategoriaForm, SalidaLineaForm,EntradaLineaForm,EntradaLineaFormSet,SalidaLineaFormSet,
    EntradaLineaFormSetMovimiento, SalidaLineaFormSetMovimiento, OrdenInsumoForm, InsumoLoteMetadataForm,
    #MovimientoLineaFormSet, 
    BodegaForm, ProveedorForm,
    EntradaForm, SalidaForm, 
    OrdenInsumoDetalleCreateFormSet, 
    OrdenInsumoDetalleEditFormSet,UnidadMedidaForm, AlertaForm
)
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, numbers
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from datetime import date,timedelta
import csv
from django.http import HttpResponse

from functools import reduce
import operator
from django.db.models import Q
from django.template.loader import render_to_string
from inventario import models


# --- Función genérica para listas con filtros, orden y paginación ---
def list_with_filters(
    request,
    base_qs,
    *,
    search_fields=None,          # lista de campos para icontains
    order_field=None,            # campo base para ordenar
    session_prefix="",           # prefijo para claves de sesión
    context_key="",              # nombre del PageObj en contexto
    full_template="",            # template completo
    partial_template="",         # template parcial
    default_per_page=20,         
    default_order="asc",         
    tie_break="id",              
    extra_context=None,          # dict extra opcional
):
    extra_context = extra_context or {}

    # --- per_page (5/10/20) con sesión por lista ---
    allowed_pp = {"5", "10", "20"}
    per_page = request.GET.get("per_page")
    if per_page in allowed_pp:
        request.session[f"per_page_{session_prefix}"] = int(per_page)
    per_page = request.session.get(f"per_page_{session_prefix}", default_per_page)

    # --- búsqueda ---
    q = (request.GET.get("q") or "").strip()

    if search_fields:
        q_objs = []
        for f in search_fields:
            try:
                # Probar acceso vía values() evita fallar por campo inexistente
                base_qs.model._meta.get_field(f.split("__")[0])
                q_objs.append(Q(**{f"{f}__icontains": q}))
            except Exception:
                # Si el primer tramo no es field directo (FK anidada), igual intentamos usarlo
                q_objs.append(Q(**{f"{f}__icontains": q}))
        if q:
            base_qs = base_qs.filter(reduce(operator.or_, q_objs))

    # --- orden ---
    allowed_order = {"asc", "desc"}
    order = request.GET.get("order")
    if order in allowed_order:
        request.session[f"order_{session_prefix}"] = order
    order = request.session.get(f"order_{session_prefix}", default_order)

    if order_field:
        ordering = order_field if order == "asc" else f"-{order_field}"
        base_qs = base_qs.order_by(ordering, tie_break)
    else:
        # Si no hay campo de orden, mantenemos el QS (pero con desempate para determinismo)
        base_qs = base_qs.order_by(tie_break)

    # --- paginación ---
    paginator = Paginator(base_qs, per_page)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # --- armar contexto común ---
    context = {
        context_key: page_obj,
        "per_page": per_page,
        "q": q,
        "order": order,
        **extra_context,
        "request": request,  # necesario para preservar GET en links dentro del partial
    }

    # --- respuesta AJAX (solo fragmento) ---
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        html = render_to_string(partial_template, context, request=request)
        return JsonResponse({"html": html})

    # --- respuesta normal (página completa) ---
    return render(request, full_template, context)

# --- DASHBOARD (ACTUALIZADO: Interactivo) ---
@login_required
def dashboard_view(request):
    # Cuentas totales
    total_insumos = Insumo.objects.filter(is_active=True).count()
    total_bodegas = Bodega.objects.filter(is_active=True).count()
    ordenes_pendientes_count = OrdenInsumo.objects.filter(estado='PENDIENTE').count()
    visitas = request.session.get('visitas', 0)
    request.session['visitas'] = visitas + 1

    # --- Listas para el Dashboard (Top 5) ---
    top_insumos = Insumo.objects.filter(is_active=True).select_related('categoria', 'unidad_medida').order_by('-created_at')[:5] 
    top_bodegas = Bodega.objects.filter(is_active=True).order_by('-created_at')[:5]
    top_ordenes = OrdenInsumo.objects.filter(estado='PENDIENTE').select_related('usuario').order_by('-fecha')[:5] 
    
    # NUEVO: Alertas (Top 5 más recientes)
    top_alertas = models.AlertaInsumo.objects.filter(is_active=True).select_related('insumo').order_by('-fecha')[:5] # MODIFICADO
    total_alertas = models.AlertaInsumo.objects.filter(is_active=True).count() # MODIFICADO

    context = {
        'total_insumos': total_insumos,
        'total_bodegas': total_bodegas,
        'ordenes_pendientes_count': ordenes_pendientes_count, # Usar el conteo correcto
        'top_insumos': top_insumos,          
        'top_bodegas': top_bodegas,          
        'top_ordenes': top_ordenes,
        'top_alertas': top_alertas,          # NUEVO
        'total_alertas': total_alertas,      # NUEVO
        'visitas': visitas,
    }
    return render(request, 'dashboard.html', context)

# heladeria/inventario/views.py (Modificando listar_alertas)

@login_required
@perfil_required(allow=("administrador", "Encargado"))
def listar_alertas(request):
    """
    Listado de alertas con filtros y paginación.
    Permite ver el historial de alertas (inactivas).
    """
    
    # NUEVO: Filtro para mostrar todas (activas e inactivas)
    mostrar_inactivas = request.GET.get("mostrar_inactivas", "0") == "1"
    
    qs = (
        models.AlertaInsumo.objects.all()
    )
    
    if not mostrar_inactivas:
        # Si no se pide el historial, se filtra por activas (comportamiento por defecto)
        qs = qs.filter(is_active=True)
    
    qs = qs.select_related("insumo").order_by("-fecha", "-created_at")

    read_only = not (
        request.user.is_superuser
        or user_has_role(request.user, "Administrador", "Encargado")
    )

    return list_with_filters(
        request,
        qs,
        search_fields=["insumo__nombre", "tipo", "mensaje"],
        order_field="fecha", 
        session_prefix="alertas",
        context_key="alertas",
        full_template="inventario/listar_alertas.html", 
        partial_template="inventario/partials/alertas_results.html", 
        default_per_page=20,
        default_order="desc", 
        tie_break="id",
        extra_context={
            "titulo": "Alertas de Inventario",
            "read_only": read_only,
            "mostrar_inactivas": mostrar_inactivas, # <--- Se pasa al contexto para el botón de historial
        },
    )
    
@login_required
@perfil_required(allow=("administrador", "Encargado"))
def crear_alerta(request):
    """Maneja la creación de una nueva alerta de inventario."""
    if not user_has_role(request.user, "Administrador", "Encargado"):
        messages.error(request, "No tienes permisos para crear alertas.")
        return redirect('inventario:listar_alertas')

    # Asumiendo que existe una clase AlertaForm en forms.py
    form = AlertaForm(request.POST or None) 

    if request.method == 'POST' and form.is_valid():
        alerta = form.save(commit=False)
        # Asumiendo que la alerta no necesita el campo usuario, si lo necesita, ajusta esta línea:
        # alerta.usuario = request.user 
        alerta.save()
        messages.success(request, "✅ Alerta creada correctamente.")
        return redirect('inventario:listar_alertas')

    # Retorno: Necesitas crear la plantilla inventario/crear_alerta.html
    return render(request, 'inventario/crear_alerta.html', {
        'form': form, 
        'titulo': 'Nueva Alerta de Inventario'
    })    
    
@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def eliminar_alerta(request, pk):
    """
    Elimina (desactiva) una alerta de inventario específica.
    """
    alerta = get_object_or_404(models.AlertaInsumo, pk=pk)

    if request.method == "POST":
        # Opción 1: Eliminación suave (Soft delete - recomendado para histórico)
        alerta.is_active = False
        alerta.save(update_fields=["is_active"])
        
        # Opción 2: Eliminación permanente
        # alerta.delete() 

        messages.success(request, f"La alerta para '{alerta.insumo.nombre}' ha sido marcada como revisada/eliminada.")
        return redirect('inventario:listar_alertas')
    
    # Si es una petición GET, se muestra la plantilla de confirmación
    return render(
        request,
        "inventario/confirmar_eliminar.html", # Reutilizamos la plantilla genérica
        {
            "titulo": f"Marcar Alerta #{alerta.id} como Revisada",
            "obj": alerta,
            "tipo": "alerta",
            "cancel_url": reverse("inventario:listar_alertas"),
        }
    )

# heladeria/inventario/views.py

# ... (Asegúrate de que estas importaciones existen al inicio del archivo)
from django.http import JsonResponse 
from django.db.models import Sum, Q, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal
# ...

# --- Función AJAX: Obtener información de stock y límites para un Insumo ---

@login_required
def get_insumo_stock_info(request, insumo_id):
    """
    Retorna en formato JSON el stock actual, stock mínimo y stock máximo 
    para un insumo dado, usado para asistir al usuario en formularios de movimiento.
    """
    try:
        # 1. Obtener el insumo y calcular el stock actual en una sola consulta
        insumo = (
            Insumo.objects.filter(id=insumo_id, is_active=True)
            .annotate(
                stock_actual=Coalesce(
                    Sum('lotes__cantidad_actual', filter=Q(lotes__is_active=True)), 
                    Decimal("0.00"), 
                    output_field=DecimalField(max_digits=10, decimal_places=2)
                )
            )
            .select_related('unidad_medida')
            .get()
        )
        
        # 2. Determinar el mensaje de ayuda (simulación de inteligencia)
        current_stock = insumo.stock_actual
        min_stock = insumo.stock_minimo
        max_stock = insumo.stock_maximo
        unidad_corto = insumo.unidad_medida.nombre_corto

        mensaje_ayuda = ""

        if current_stock < min_stock:
            cantidad_sugerida = max_stock - current_stock
            mensaje_ayuda = (
                f"🚨 **Stock Bajo**: {current_stock:.2f} {unidad_corto}. "
                f"Sugerencia: Entrar **{cantidad_sugerida:.2f}** {unidad_corto} para alcanzar el máximo ({max_stock:.2f})."
            )
        elif current_stock > max_stock:
            cantidad_sugerida = current_stock - max_stock
            mensaje_ayuda = (
                f"⚠️ **Stock Excesivo**: {current_stock:.2f} {unidad_corto}. "
                f"Sugerencia: Salir **{cantidad_sugerida:.2f}** {unidad_corto} para volver al máximo ({max_stock:.2f})."
            )
        else:
            mensaje_ayuda = f"✅ **Stock OK**: {current_stock:.2f} {unidad_corto}. Rango óptimo: {min_stock:.2f} - {max_stock:.2f}."


        return JsonResponse({
            'success': True,
            'stock_actual': float(current_stock),
            'stock_minimo': float(min_stock),
            'stock_maximo': float(max_stock),
            'unidad_medida': unidad_corto,
            'mensaje_ayuda': mensaje_ayuda,
        })

    except Insumo.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Insumo no encontrado o inactivo.'}, status=404)
    except Exception as e:
        # Manejo genérico de errores (ej. si falta el insumo_id en la URL)
        return JsonResponse({'success': False, 'message': f'Error interno: {str(e)}'}, status=500)

# --- CRUD INSUMOS ---
@login_required
@perfil_required(
    allow=("administrador", "Encargado"),
    readonly_for=("Bodeguero",)             
)
def listar_insumos(request):
    qs = (
        Insumo.objects.filter(is_active=True)
        .annotate(stock_actual=Coalesce(Sum('lotes__cantidad_actual'), 0, output_field=DecimalField()))
        .select_related('categoria', 'unidad_medida') # Optimizada
    )

    allowed_sort = {"nombre", "categoria", "stock", "unidad"}
    sort = request.GET.get("sort")
    if sort not in allowed_sort:
        sort = request.session.get("sort_insumos", "nombre")
    if sort not in allowed_sort:
        sort = "nombre"
    request.session["sort_insumos"] = sort

    sort_map = {
        "nombre": "nombre",
        "categoria": "categoria__nombre",
        "stock": "stock_actual",
        "unidad": "unidad_medida__nombre_largo", # Ordena por nombre de la unidad
    }
    read_only = not (request.user.is_superuser or user_has_role(request.user, "Administrador", "Admin", "Encargado"))
    
    return list_with_filters(
        request,
        qs,
        # Búsqueda: ahora en el campo de texto de la unidad
        search_fields=["nombre", "categoria__nombre", "unidad_medida__nombre_largo"], 
        order_field=sort_map[sort],
        session_prefix="insumos",
        context_key="insumos",
        full_template="inventario/listar_insumos.html",
        partial_template="inventario/partials/insumos_results.html",
        default_per_page=10, # Este es solo el fallback inicial. list_with_filters usa la sesión o el valor GET.
        default_order="asc",
        tie_break="id",
        extra_context={"read_only": read_only,"titulo": "Listado de Insumos", "sort": sort},
    )

# --- CREAR INSUMO (Actualizado: Pasa UnidadMedidaForm) ---
@login_required
@perfil_required(allow=("administrador", "Encargado"))
def crear_insumo(request):
    if not user_has_role(request.user, "Administrador", "Encargado"):
        messages.error(request, "No tienes permisos para crear insumos.")
        return redirect('inventario:listar_insumos')

    form = InsumoForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "✅ Insumo creado correctamente.")
        return redirect('inventario:listar_insumos')

    return render(request, 'inventario/crear_insumo.html', {
        'form': form, 
        'titulo': 'Nuevo Insumo',
        'UnidadMedidaForm': UnidadMedidaForm() # <--- PASAMOS EL FORMULARIO PARA EL MODAL
    })


# --- EDITAR INSUMO (Actualizado: Pasa UnidadMedidaForm) ---
@login_required
@perfil_required(allow=("administrador", "Encargado"))
def editar_insumo(request, insumo_id):
    if not user_has_role(request.user, "Administrador", "Encargado"):
        messages.error(request, "No tienes permisos para editar insumos.")
        return redirect('inventario:listar_insumos')

    insumo = get_object_or_404(Insumo, id=insumo_id)
    form = InsumoForm(request.POST or None, instance=insumo)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f"📝 Insumo '{insumo.nombre}' actualizado.")
        return redirect('inventario:listar_insumos')

    return render(request, 'inventario/editar_insumo.html', {
        'form': form, 
        'titulo': f'Editar {insumo.nombre}',
        'UnidadMedidaForm': UnidadMedidaForm() # <--- PASAMOS EL FORMULARIO PARA EL MODAL
    })


# --- ELIMINAR INSUMO (Actualizado: Stock Cero) ---
@login_required
@perfil_required(allow=("administrador", "Encargado"))
def eliminar_insumo(request, insumo_id):
    # 1. Obtener el Insumo y calcular el stock total de sus lotes activos
    insumo = get_object_or_404(
        Insumo.objects.annotate(
            stock_total=Coalesce(
                Sum('lotes__cantidad_actual', filter=Q(lotes__is_active=True)), 
                Decimal("0.00"), 
                output_field=DecimalField()
            )
        ), 
        id=insumo_id
    )
    
    stock_actual = insumo.stock_total

    if request.method == "POST":
        # 2. Regla: NO se puede eliminar si el stock actual es mayor que cero.
        if stock_actual > Decimal("0.00"):
            messages.error(
                request, 
                f"🚫 No se puede eliminar el insumo '{insumo.nombre}' porque tiene stock activo ({stock_actual:.2f})."
            )
            return redirect("inventario:listar_insumos")
            
        # 3. Procede con la eliminación (desactivación) si el stock es cero
        insumo.is_active = False
        insumo.save(update_fields=["is_active"])
        messages.success(request, f"🗑️ El insumo '{insumo.nombre}' fue eliminado.")
        return redirect("inventario:listar_insumos")

    return render(
        request,
        "inventario/confirmar_eliminar.html", # Ruta corregida previamente
        {"titulo": f"Eliminar Insumo «{insumo.nombre}»", "cancel_url": reverse("inventario:listar_insumos"), "insumo": insumo}
    )

# --- VISTA AJAX (NUEVA: Creación de Unidad de Medida) ---
@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
@require_POST
def crear_unidad_medida_ajax(request):
    """Crea una nueva unidad de medida y devuelve sus datos en JSON."""
    form = UnidadMedidaForm(request.POST)
    
    if form.is_valid():
        nueva_unidad = form.save()
        return JsonResponse({
            'success': True,
            'id': nueva_unidad.id,
            'text': str(nueva_unidad) # Retorna "Kilogramos (KG)"
        }, status=201)
    
    # Si hay errores de validación, retornamos JSON con los errores
    errors = json.loads(form.errors.as_json())
    return JsonResponse({
        'success': False,
        'errors': errors
    }, status=400)

@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
@require_POST
def editar_unidad_medida_ajax(request, pk):
    """Edita una unidad de medida existente y devuelve sus datos en JSON."""
    
    # 1. Obtener la instancia existente
    try:
        unidad = UnidadMedida.objects.get(pk=pk)
    except UnidadMedida.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Unidad no encontrada.'}, status=404)

    # 2. Inicializar el formulario con la data POST y la instancia existente
    form = UnidadMedidaForm(request.POST, instance=unidad)

    if form.is_valid():
        unidad_actualizada = form.save()
        return JsonResponse({
            'success': True,
            'id': unidad_actualizada.id,
            'text': str(unidad_actualizada), 
            'message': f'Unidad "{str(unidad_actualizada)}" actualizada con éxito.'
        }, status=200)
    
    # Si hay errores de validación, retornamos JSON con los errores
    import json
    errors = json.loads(form.errors.as_json())
    return JsonResponse({
        'success': False,
        'errors': errors,
        'message': 'Error de validación al actualizar la unidad.'
    }, status=400)
    
@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
@require_POST
def eliminar_unidad_medida_ajax(request, pk):
    """Elimina una unidad de medida por PK tras verificar que no esté en uso."""
    try:
        unidad = UnidadMedida.objects.get(pk=pk)
    except UnidadMedida.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Unidad no encontrada.'}, status=404)

    # 1. Validación de dependencia: ¿Está la unidad en uso por algún Insumo?
    # Importante: El modelo Insumo ahora tiene un ForeignKey a UnidadMedida
    if unidad.insumos_medidos.exists():
        return JsonResponse({
            'success': False,
            'message': f'La unidad "{str(unidad)}" está en uso y no puede eliminarse.'
        }, status=400)

    # 2. Eliminación (Desactivación)
    unidad.is_active = False # Asumo que usas desactivación suave (soft delete)
    unidad.save(update_fields=['is_active'])
    
    # Si quieres eliminarla permanentemente:
    # unidad.delete() 

    return JsonResponse({'success': True, 'message': f'Unidad "{str(unidad)}" eliminada.'}, status=200)

# --- exportar LOTES DE INSUMO ---
@login_required
@perfil_required(allow=("administrador", "Encargado"))
def exportar_lotes(request):
    """
    Exporta la lista de lotes de insumos a Excel o PDF, respetando filtros y orden.
    Soporta reporte especial de 'Próximos a vencer' usando ?proximos=1.
    """
    # 1. Preparación del QuerySet (Similar a listar_insumos_lote)
    qs = (
        InsumoLote.objects.filter(is_active=True)
        .select_related("insumo", "bodega", "proveedor")  # ⬅️ ahora incluye proveedor
        .annotate(
            cant_act=Coalesce(F("cantidad_actual"), 0, output_field=DecimalField()),
            cant_ini=Coalesce(F("cantidad_inicial"), 0, output_field=DecimalField()),
        )
    )

    hoy = date.today()
    dias_proximos = 14  # rango de días para considerar "próximos a vencer"

    # ¿Es un reporte solo de próximos a vencer?
    solo_proximos = (request.GET.get("proximos") == "1")

    if solo_proximos:
        limite = hoy + timedelta(days=dias_proximos)
        qs = qs.filter(
            fecha_expiracion__isnull=False,
            fecha_expiracion__gte=hoy,
            fecha_expiracion__lte=limite,
        )

    # 🔹 NUEVO: filtro por proveedor (igual que en listar_insumos_lote)
    proveedor_id = request.GET.get("proveedor")
    if proveedor_id:
        qs = qs.filter(proveedor_id=proveedor_id)

    # Filtros de búsqueda (q)
    q = (request.GET.get("q") or "").strip()
    # también buscamos por proveedor
    search_fields = ["insumo__nombre", "bodega__nombre", "proveedor__nombre_empresa"]
    if q:
        q_objs = []
        for f in search_fields:
            q_objs.append(Q(**{f"{f}__icontains": q}))
        qs = qs.filter(reduce(operator.or_, q_objs))
    
    # Orden (sort y order)
    allowed_sort = {"insumo", "bodega", "fingreso", "fexpira", "cact", "cini"}
    sort = request.GET.get("sort")
    if sort not in allowed_sort:
        # La exportación no depende de la sesión, solo de la URL
        sort = "insumo" 

    sort_map = {
        "insumo":   "insumo__nombre",
        "bodega":   "bodega__nombre",
        "fingreso": "fecha_ingreso",
        "fexpira":  "fecha_expiracion",
        "cact":     "cant_act",
        "cini":     "cant_ini",
    }
    
    order = request.GET.get("order")
    allowed_order = {"asc", "desc"}
    if order not in allowed_order:
        order = "asc"
        
    tie_break = "id"
    ordering = sort_map[sort] if order == "asc" else f"-{sort_map[sort]}"
    qs = qs.order_by(ordering, tie_break)
    
    # Obtener todos los datos sin paginación
    lotes = qs.all()
    
    # 2. Lógica de Exportación
    exportar = request.GET.get("exportar")

    # Títulos según si es reporte general o de próximos a vencer
    if solo_proximos:
        titulo_reporte = "Reporte de Lotes de Insumos Próximos a Vencer"
        sufijo_nombre = "lotes_proximos"
    else:
        titulo_reporte = "Reporte de Lotes de Insumos"
        sufijo_nombre = "lotes"

    if exportar == "excel":
        # --- Exportar a Excel (openpyxl) ---
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{sufijo_nombre}_{hoy.isoformat()}.xlsx"'
        )

        # Workbook y Worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Lotes de Insumos"

        # Título (ahora hay 8 columnas: A–H)
        ws.merge_cells("A1:H1")
        ws.cell(row=1, column=1, value=titulo_reporte).font = Font(size=14, bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        
        # 🔹 Cabeceras con Proveedor
        headers = [
            "Insumo",
            "Bodega",
            "Proveedor",
            "Fecha de Ingreso",
            "Fecha de Expiración",
            "Cantidad Inicial",
            "Cantidad Actual",
            "ID Lote",
        ]
        ws.append(headers)
        
        # Estilo para encabezados
        header_row = ws[2]
        header_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")  # Verde claro
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill

        # Datos
        row_num = 3
        for i, lote in enumerate(lotes):
            # Alternar color de fondo de las filas
            bg_color = 'E6E6FA' if i % 2 == 0 else 'FFFFFF'  # Lavanda pálida / Blanco
            
            data = [
                lote.insumo.nombre,
                lote.bodega.nombre,
                lote.proveedor.nombre_empresa if lote.proveedor else "",  # ⬅️ proveedor
                lote.fecha_ingreso.strftime("%Y-%m-%d") if lote.fecha_ingreso else "",
                lote.fecha_expiracion.strftime("%Y-%m-%d") if lote.fecha_expiracion else "N/A",
                float(lote.cantidad_inicial),
                float(lote.cantidad_actual),
                lote.id,
            ]
            ws.append(data)
            
            # Aplicar estilo de fondo y formato de números
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            for col_idx in range(1, len(data) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = fill
                # Formato de números para las cantidades
                # ahora las cantidades están en columnas 6 y 7 (antes 5 y 6)
                if col_idx in [6, 7]:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            row_num += 1

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[1].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 4)
            # Ajustamos un poco las columnas de fechas/cantidades
            if column in ('C', 'D', 'E', 'F', 'G', 'H') and adjusted_width < 15:
                adjusted_width = 15
            ws.column_dimensions[column].width = adjusted_width
            
        buffer = BytesIO()
        wb.save(buffer)
        response.write(buffer.getvalue())
        return response

    elif exportar == "pdf":
        # --- Exportar a PDF (reportlab) ---
        bio = BytesIO()
        doc = SimpleDocTemplate(
            bio,
            pagesize=landscape(A4),
            leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20,
        )
        
        story = []
        styles = getSampleStyleSheet()

        # Título
        titulo = Paragraph(f"<b>{titulo_reporte}</b>", styles["h1"])
        story.append(titulo)
        story.append(Spacer(1, 12))

        # Datos de la tabla
        data = [
            ["Insumo", "Bodega", "Proveedor", "F. Ingreso", "F. Exp.", "Cant. Inicial", "Cant. Actual", "ID Lote"]
        ]
        
        for lote in lotes:
            data.append([
                lote.insumo.nombre,
                lote.bodega.nombre,
                lote.proveedor.nombre_empresa if lote.proveedor else "",
                lote.fecha_ingreso.strftime("%Y-%m-%d") if lote.fecha_ingreso else "",
                lote.fecha_expiracion.strftime("%Y-%m-%d") if lote.fecha_expiracion else "N/A",
                f"{lote.cantidad_inicial:,.2f}",
                f"{lote.cantidad_actual:,.2f}",
                str(lote.id),
            ])
            
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A86E8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 1), (2, -1), 'LEFT'),
            ('ALIGN', (3, 1), (-2, -1), 'CENTER'),
            ('ALIGN', (-1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F5F5F5'), colors.white]),
        ]
        
        # A4 apaisado — 8 columnas
        t = Table(
            data,
            colWidths=[200, 100, 140, 70, 70, 90, 90, 50]
        )
        t.setStyle(TableStyle(table_style))
        story.append(t)
        
        doc.build(story)
        pdf_value = bio.getvalue()
        bio.close()
        response = HttpResponse(pdf_value, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="{sufijo_nombre}_{hoy.isoformat()}.pdf"'
        )
        return response

    return HttpResponseBadRequest("Método de exportación no válido.")

# --- LISTAR LOTES DE INSUMO ---
@login_required
@perfil_required(allow=("administrador", "Encargado"), readonly_for=("Bodeguero",))
def listar_insumos_lote(request):
    hoy = date.today()

    qs = (
        InsumoLote.objects.filter(is_active=True)
        .select_related("insumo", "bodega", "proveedor")  # ⬅️ añadimos proveedor
        .annotate(
            cant_act=Coalesce(F("cantidad_actual"), 0, output_field=DecimalField()),
            cant_ini=Coalesce(F("cantidad_inicial"), 0, output_field=DecimalField()),
        )
    )

    # -------- filtro de vencimiento --------
    vencimiento = request.GET.get("vencimiento")  # None o "proximos"
    default_dias = 30
    try:
        dias = int(request.GET.get("dias", default_dias))
        if dias <= 0 or dias > 365:
            dias = default_dias
    except ValueError:
        dias = default_dias

    if vencimiento == "proximos":
        limite = hoy + timedelta(days=dias)
        qs = qs.filter(
            fecha_expiracion__isnull=False,
            fecha_expiracion__gte=hoy,
            fecha_expiracion__lte=limite,
        )

    # -------- NUEVO: filtro por proveedor --------
    proveedor_id = request.GET.get("proveedor")   # ⬅️ aquí nace proveedor_sel
    if proveedor_id:
        qs = qs.filter(proveedor_id=proveedor_id)

    # sort permitido
    allowed_sort = {"insumo", "bodega", "fingreso", "fexpira", "cact", "cini"}
    sort = request.GET.get("sort")
    if sort not in allowed_sort:
        sort = request.session.get("sort_lotes", "insumo")
    if sort not in allowed_sort:
        sort = "insumo"
    request.session["sort_lotes"] = sort

    sort_map = {
        "insumo":  "insumo__nombre",
        "bodega":  "bodega__nombre",
        "fingreso":"fecha_ingreso",
        "fexpira": "fecha_expiracion",
        "cact":    "cant_act",
        "cini":    "cant_ini",
    }

    read_only = not (
        request.user.is_superuser or
        user_has_role(request.user, "Administrador", "Admin", "Encargado")
    )

    return list_with_filters(
        request,
        qs,
        # ⬅️ ahora la búsqueda de texto también mira proveedor
        search_fields=["insumo__nombre", "bodega__nombre", "proveedor__nombre_empresa"],
        order_field=sort_map[sort],
        session_prefix="lotes",
        context_key="lotes",
        full_template="inventario/listar_insumo_lote.html",
        partial_template="inventario/partials/insumo_lote_results.html",
        default_per_page=10,
        default_order="asc",
        tie_break="id",
        extra_context={
            "read_only": read_only,
            "titulo": "Lotes de Insumos",
            "sort": sort,
            "today": hoy,
            "filtro_vencimiento": vencimiento or "todos",
            "dias_proximos": dias,

            # 🔥 de aquí lo “trae” el template
            "proveedores": Proveedor.objects.all().order_by("nombre_empresa"),
            "proveedor_sel": proveedor_id,   # ⬅️ eso es lo que estás usando en el partial
        },
    )

@login_required
@perfil_required(allow=("administrador", "Encargado"))
def crear_lote(request):
    """
    Redirige a registrar_entrada ya que los lotes deben crearse con stock inicial.
    """
    messages.warning(request, "Los lotes de insumos deben crearse a través del registro de una Entrada para iniciar el stock correctamente.")
    return redirect(reverse('inventario:registrar_entrada'))


@login_required
@perfil_required(allow=("administrador", "Encargado"))
def editar_lote(request, pk):
    lote = get_object_or_404(models.InsumoLote.objects.select_related('insumo'), pk=pk)
    
    # ⚠️ Regla: Bloquear edición si la cantidad actual es cero o negativa
    # Esto evita problemas al modificar una fecha de expiración de un lote "agotado" si se quiere eliminar después.
    if lote.cantidad_actual <= Decimal("0.00") and lote.is_active:
        messages.error(request, f"No se pueden editar los metadatos del Lote #{pk} porque su stock es cero y ya fue utilizado.")
        return redirect(reverse('inventario:listar_lotes'))
        
    form = InsumoLoteMetadataForm(request.POST or None, instance=lote)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f"📝 Metadatos del Lote #{pk} de {lote.insumo.nombre} actualizados.")
        return redirect(reverse('inventario:listar_lotes')) 

    # Se usará una plantilla genérica llamada 'lote_form.html'
    return render(request, 'inventario/lote_form.html', {
        'form': form,
        'titulo': f'Editar Lote #{pk}',
        'lote': lote,
    })


@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def eliminar_lote(request, pk):
    lote = get_object_or_404(models.InsumoLote.objects.select_related('insumo'), pk=pk)
    
    # Regla: No se puede eliminar si tiene stock actual > 0
    if lote.cantidad_actual > Decimal("0.00"):
        messages.error(request, f"🚫 No se puede eliminar el Lote #{pk} porque tiene stock activo ({lote.cantidad_actual}). Debe registrar una salida para agotarlo primero.")
        return redirect(reverse('inventario:listar_lotes'))

    if request.method == "POST":
        lote.is_active = False # Soft delete
        lote.save(update_fields=["is_active"])
        messages.success(request, f"🗑️ Lote #{pk} de {lote.insumo.nombre} desactivado.")
        return redirect(reverse('inventario:listar_lotes'))

    # Reutilizar plantilla de confirmación genérica
    return render(
        request,
        "inventario/confirmar_eliminar.html",
        {
            "titulo": f"Eliminar Lote «{lote.id} - {lote.insumo.nombre}»",
            "obj": lote,
            "tipo": "lote",
            "cancel_url": reverse("inventario:listar_lotes"),
            "extra_message": f"El stock actual es {lote.cantidad_actual}. Solo puede eliminarse si es cero. Al confirmar, se marcará como inactivo."
        }
    )
    
@login_required
@perfil_required(allow=("administrador", "Encargado", "Bodeguero"))
def ver_detalle_lote(request, pk):
    lote = get_object_or_404(
        models.InsumoLote.objects.select_related(
            'insumo__unidad_medida', 
            'bodega', 
            'proveedor', 
            'usuario'
        ), 
        pk=pk
    )
    
    # Obtener los movimientos asociados para el detalle
    entradas = models.Entrada.objects.filter(insumo_lote=lote).order_by('-fecha')
    salidas = models.Salida.objects.filter(insumo_lote=lote).order_by('-fecha_generada')

    return render(request, 'inventario/ver_detalle_lote.html', {
        'titulo': f'Detalle Lote #{pk}',
        'lote': lote,
        'entradas': entradas,
        'salidas': salidas,
    })

# --- CATEGORÍAS ---
@login_required
@perfil_required(allow=("administrador", "Encargado"), readonly_for=("Bodeguero",))
def listar_categorias(request):
    allowed_pp = {"5", "10", "20"}
    per_page_get = request.GET.get("per_page")
    if per_page_get in allowed_pp:
        request.session["per_page_categorias"] = int(per_page_get)
    per_page = request.session.get("per_page_categorias", 10)

    order_get = request.GET.get("order")
    if order_get in ("asc", "desc"):
        request.session["order_categorias"] = order_get
    order = request.session.get("order_categorias", "asc")

    q = (request.GET.get("q") or "").strip()

    # CORRECCIÓN CLAVE: FILTRAR POR is_active=True
    categorias = Categoria.objects.filter(is_active=True).order_by(
        "nombre" if order == "asc" else "-nombre"
    )
    if q:
        categorias = categorias.filter(
            Q(nombre__icontains=q) | Q(descripcion__icontains=q)
        )

    paginator = Paginator(categorias, per_page)
    page_number = request.GET.get("page")
    categorias_page = paginator.get_page(page_number)

    read_only = not (
        request.user.is_superuser
        # 🔥 CORREGIDO: Incluye Encargado en la lista de roles con permisos de escritura
        or user_has_role(request.user, "Administrador", "Encargado")
    )

    context = {
        "categorias": categorias_page,
        "per_page": per_page,
        "q": q,
        "order": order,
        "read_only": read_only,
    }

    # ⚡ AJAX -> JSON
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        html = render_to_string(
            "inventario/partials/categorias_results.html",
            context,
            request=request,
        )
        return JsonResponse({"html": html})

    # render normal
    return render(
        request,
        "inventario/listar_categorias.html",
        context,
    )

@login_required
@perfil_required(allow=("administrador", "Encargado"))
def crear_categoria(request):
    if not user_has_role(request.user, "administrador","encargado" ):
        messages.error(request, "No tienes permisos.")
        return redirect('inventario:listar_categorias')

    form = CategoriaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "✅ Categoría creada.")
        return redirect('inventario:listar_categorias')

    return render(request, 'inventario/crear_categoria.html', {'form': form, 'titulo': 'Nueva Categoría'})

@login_required
@perfil_required(allow=("administrador", "Encargado"))
def editar_categoria(request, categoria_id):
    if not user_has_role(request.user, "Administrador","encargado"):
        messages.error(request, "No tienes permisos para editar categorías.")
        return redirect('inventario:listar_categorias')

    categoria = get_object_or_404(Categoria, id=categoria_id)
    form = CategoriaForm(request.POST or None, instance=categoria)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f"Categoría '{categoria.nombre}' actualizada.")
        return redirect('inventario:listar_categorias')

    context = {'form': form, 'titulo': f'Editar Categoría: {categoria.nombre}'}
    return render(request, 'inventario/editar_categoria.html', context)

@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def eliminar_categoria(request, categoria_id):
    """Elimina (soft delete) una categoría, solo si no tiene insumos activos."""
    
    # Utilizamos models.Categoria para asegurar la referencia correcta
    categoria = get_object_or_404(models.Categoria, pk=categoria_id) 

    if request.method == "POST":
        
        # 1. Verificar si hay insumos activos relacionados (usando la relación inversa 'insumos')
        # Esto verifica si existe ALGÚN insumo activo que dependa de esta categoría.
        if categoria.insumos.filter(is_active=True).exists():
            messages.error(
                request,
                f"🚫 No se puede eliminar la categoría '{categoria.nombre}' porque tiene insumos activos asociados. Desactive primero los insumos vinculados."
            )
            return redirect('inventario:listar_categorias')

        # 2. Realizar la eliminación suave (soft delete)
        categoria.is_active = False
        categoria.save(update_fields=['is_active'])
        messages.success(request, f"🗑️ Categoría '{categoria.nombre}' desactivada correctamente.")
        return redirect('inventario:listar_categorias')

    # Si es GET, se dirige a la plantilla de confirmación
    return render(
        request,
        "inventario/confirmar_eliminar.html", # Usamos la plantilla genérica
        {
            "titulo": f"Eliminar Categoría «{categoria.nombre}»",
            "obj": categoria,
            "tipo": "categoría",
            "cancel_url": reverse("inventario:listar_categorias"),
        }
    )

'''
# --- MOVIMIENTOS UNIFICADOS ---
@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def registrar_movimiento(request):
    """Registrar entradas/salidas (pueden o no pertenecer a una orden)."""
    if not user_has_role(request.user, "Administrador", "Encargado"):
        messages.error(request, "No tienes permisos para registrar movimientos.")
        return redirect('inventario:listar_movimientos')

    orden_obj = None
    # Si venimos desde una orden (?orden=ID)
    if request.method == "GET":
        orden_id = request.GET.get("orden")
        if orden_id:
            orden_obj = get_object_or_404(OrdenInsumo, id=orden_id)
    else:
        orden_id = request.POST.get("orden_id")
        if orden_id:
            orden_obj = get_object_or_404(OrdenInsumo, id=orden_id)

    if request.method == "POST":
        formset = MovimientoLineaFormSet(request.POST)
        if not formset.is_valid():
            messages.error(request, "Revisa los errores en el formulario.")
            return render(request, "inventario/registrar_movimiento.html", {
                "formset": formset,
                "titulo": "Registrar movimiento",
                "orden": orden_obj,
            })

        lineas = [f.cleaned_data for f in formset if getattr(f, "cleaned_data", None) and not f.cleaned_data.get("DELETE")]
        if not lineas:
            messages.warning(request, "No se ingresó ninguna línea válida.")
            return render(request, "inventario/registrar_movimiento.html", {
                "formset": formset,
                "titulo": "Registrar movimiento",
                "orden": orden_obj,
            })

        # Procesar líneas
        for cd in lineas:
            tipo = cd["tipo"]
            insumo = cd["insumo"]
            ubicacion = cd["ubicacion"]
            cantidad = cd["cantidad"]
            fecha = cd["fecha"]
            obs = cd.get("observaciones", "")
            crear_nuevo = cd.get("crear_nuevo_lote")
            lote = cd.get("insumo_lote")
            fecha_exp = cd.get("fecha_expiracion")

            detalle = None
            if orden_obj:
                # Crear un detalle si no existe
                detalle = OrdenInsumoDetalle.objects.create(
                    orden_insumo=orden_obj,
                    insumo=insumo,
                    cantidad_solicitada=cantidad,
                )

            if tipo == "ENTRADA":
                if crear_nuevo:
                    lote = InsumoLote.objects.create(
                        insumo=insumo,
                        bodega=ubicacion.bodega,
                        fecha_ingreso=fecha,
                        fecha_expiracion=fecha_exp,
                        cantidad_inicial=cantidad,
                        cantidad_actual=Decimal("0.00"), # <-- CORRECCIÓN DOBLE CONTEO
                        usuario=request.user,
                    )
                elif not lote:
                    messages.error(request, "Debes seleccionar o crear un lote para ENTRADA.")
                    raise transaction.TransactionManagementError("Entrada sin lote")

                entrada = Entrada.objects.create(
                    insumo=insumo, insumo_lote=lote, ubicacion=ubicacion,
                    cantidad=cantidad, fecha=fecha, usuario=request.user,
                    observaciones=obs, orden=orden_obj, detalle=detalle
                )
                lote.cantidad_actual += cantidad
                lote.save(update_fields=["cantidad_actual"])

                if detalle:
                    detalle.cantidad_atendida += cantidad
                    detalle.save(update_fields=["cantidad_atendida"])

            elif tipo == "SALIDA":     
                if crear_nuevo:
                    messages.error(request, "No aplica 'Crear nuevo lote' para SALIDA.")
                    raise transaction.TransactionManagementError("Salida con crear_nuevo_lote")
                if not lote:
                    messages.error(request, "Para una SALIDA debes indicar el lote.")
                    raise transaction.TransactionManagementError("Salida sin lote")

                cant = min(cantidad, lote.cantidad_actual or Decimal("0"))
                Salida.objects.create(
                    insumo=insumo, insumo_lote=lote, ubicacion=ubicacion,
                    cantidad=cant, fecha_generada=fecha, usuario=request.user,
                    tipo="USO_PRODUCCION", observaciones=obs,
                    orden=orden_obj, detalle=detalle
                )
                lote.cantidad_actual -= cant # <-- Resta de stock (la lógica ya era correcta)
                lote.save(update_fields=["cantidad_actual"])

        if orden_obj:
            orden_obj.recalc_estado()

        messages.success(request, "Movimiento registrado correctamente.")
        return redirect('inventario:listar_movimientos')

    # Bloque GET, modificado para precargar insumo y tipo desde listar_insumos
    formset = MovimientoLineaFormSet()
    
    # Manejar precarga si vienen insumo_id y tipo (desde listar_insumos)
    if request.method == "GET":
        insumo_id = request.GET.get("insumo_id")
        tipo_movimiento = request.GET.get("tipo") # ENTRADA o SALIDA
        
        insumo_obj = None
        if insumo_id:
            try:
                insumo_obj = get_object_or_404(Insumo, id=insumo_id)
            except Exception:
                messages.warning(request, "El insumo solicitado no es válido.")

        # Si hay un insumo válido, precargamos el formset
        if insumo_obj and tipo_movimiento in ("ENTRADA", "SALIDA"):
            initial_data = [{
                'insumo': insumo_obj.id,
                'tipo': tipo_movimiento,
                # Dejamos cantidad y ubicación vacíos para que el usuario los complete
            }]
            formset = MovimientoLineaFormSet(initial=initial_data)

    return render(request, "inventario/registrar_movimiento.html", {
        "formset": formset,
        "titulo": "Registrar movimiento",
        "orden": orden_obj,
    })'''

@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def registrar_entrada(request):
    if not user_has_role(request.user, "Administrador", "Encargado"):
        messages.error(request, "No tienes permisos para registrar entradas.")
        return redirect('inventario:listar_movimientos')

    initial_data = []
    
    # --- LÓGICA DE PRE-CARGA (Insumo ID - SÓLO para registros directos) ---
    if request.GET.get("insumo_id"):
        insumo_id = request.GET.get("insumo_id")
        try:
            insumo_obj = models.Insumo.objects.get(id=insumo_id)
            initial_data.append({
                'insumo': insumo_obj.id,
                'cantidad': 1,
            })
        except models.Insumo.DoesNotExist:
            messages.warning(request, "El insumo solicitado no es válido.")

    if request.method == "POST":
        formset = EntradaLineaFormSet(request.POST)
        
        # Ignorar orden_id ya que esta vista es solo para registros directos
        orden_obj_post = None 
        
        # ... (Mantener la lógica de POST igual que la versión anterior de Entrada) ...

        if formset.is_valid():
            lineas = [f.cleaned_data for f in formset if getattr(f, "cleaned_data", None) and not f.cleaned_data.get("DELETE")]
            if not lineas:
                messages.warning(request, "No se ingresó ninguna línea válida para registrar.")
                return render(request, "inventario/registrar_entrada.html", {
                    "formset": formset,
                    "titulo": "Registrar Entrada de Inventario",
                })
            
            for cd in lineas:
                insumo = cd.get("insumo")
                ubicacion = cd.get("ubicacion")
                cantidad = cd.get("cantidad")
                fecha = cd.get("fecha")
                proveedor = cd.get("proveedor")
                fecha_exp = cd.get("fecha_expiracion")
                
                if not all([insumo, ubicacion, cantidad, fecha, proveedor, fecha_exp]):
                    messages.warning(request, "Datos incompletos en una línea de entrada. Se ignoró esta línea.")
                    continue
                
                # En esta vista no hay detalle/orden
                detalle_obj = None

                lote = models.InsumoLote.objects.create(
                    insumo=insumo,
                    bodega=ubicacion.bodega,
                    proveedor=proveedor,
                    fecha_ingreso=fecha,
                    fecha_expiracion=fecha_exp,
                    cantidad_inicial=cantidad,
                    cantidad_actual=Decimal("0.00"),
                    usuario=request.user,
                )

                models.Entrada.objects.create(
                    insumo=insumo, insumo_lote=lote, ubicacion=ubicacion,
                    cantidad=cantidad, fecha=fecha, usuario=request.user,
                    observaciones=cd.get("observaciones", ""),
                    orden=orden_obj_post,     
                    detalle=detalle_obj,      
                )
                
                lote.cantidad_actual += cantidad
                lote.save(update_fields=["cantidad_actual"])
                
                check_and_create_stock_alerts(insumo)

            messages.success(request, "✅ Entrada(s) registrada(s) correctamente.")
            return redirect('inventario:listar_movimientos')

        messages.error(request, "Revisa los errores en el formulario antes de continuar.")

    else:
        # Si hay precarga (insumo), no crear línea extra vacía
        if initial_data:
            DynamicFormSet = formset_factory(EntradaLineaForm, extra=0, can_delete=True)
        else:
            DynamicFormSet = EntradaLineaFormSet
        formset = DynamicFormSet(initial=initial_data if initial_data else None)

    return render(request, "inventario/registrar_entrada.html", {
        "formset": formset,
        "titulo": "Registrar Entrada de Inventario",
        "orden": None, # <--- Siempre None en la vista genérica
    })



# --- REGISTRAR MOVIMIENTO (SALIDA DEDICADA) ---
@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def registrar_salida(request):
    if not user_has_role(request.user, "Administrador", "Encargado"):
        messages.error(request, "No tienes permisos para registrar salidas.")
        return redirect('inventario:listar_movimientos')

    orden_obj = None
    initial_data = []

    # --- LÓGICA DE PRE-CARGA (Orden ID) ---
    orden_id = request.GET.get("orden")
    if orden_id:
        # 1. Obtener la orden
        orden_obj = get_object_or_404(models.OrdenInsumo, id=orden_id)
        
        # 2. Filtrar solo los detalles pendientes (solicitada > atendida)
        # Se usa select_related('insumo') para cargar el objeto Insumo y obtener el ID fácilmente.
        detalles_pendientes = orden_obj.detalles.filter(
            cantidad_solicitada__gt=F('cantidad_atendida')
        ).select_related('insumo')

        for detalle in detalles_pendientes:
            
            # Calcular la cantidad que falta por atender
            cantidad_restante = detalle.cantidad_solicitada - detalle.cantidad_atendida
            
            # ✅ CÓDIGO CORREGIDO: Usamos la cantidad restante (Decimal) directamente para el initial data
            cantidad_inicial = cantidad_restante

            # 4. Añadir solo si la cantidad inicial es positiva
            if cantidad_inicial > Decimal("0.00"): # Se usa Decimal para comparar
                initial_data.append({
                    'insumo': detalle.insumo.id, # <-- ID del Insumo
                    'cantidad': cantidad_inicial, # <-- Se pasa como Decimal
                })
    
    # --- LÓGICA DE PRE-CARGA (Insumo ID - Si NO HAY Orden) ---
    elif request.GET.get("insumo_id"):
        insumo_id = request.GET.get("insumo_id")
        try:
            insumo_obj = models.Insumo.objects.get(id=insumo_id)
            initial_data.append({
                'insumo': insumo_obj.id,
                'cantidad': 1, 
            })
        except models.Insumo.DoesNotExist:
            messages.warning(request, "El insumo solicitado no es válido.")

    if request.method == "POST":
        formset = SalidaLineaFormSet(request.POST)
        
        # Recuperar orden si se envió el ID oculto (por si no se usó GET)
        orden_id_post = request.POST.get("orden_id")
        orden_obj_post = orden_obj
        if not orden_obj_post and orden_id_post:
             orden_obj_post = get_object_or_404(models.OrdenInsumo, id=orden_id_post)


        if formset.is_valid():
            lineas = [f.cleaned_data for f in formset if getattr(f, "cleaned_data", None) and not f.cleaned_data.get("DELETE")]
            
            # CRÍTICO: Bloquear el éxito si no hay líneas válidas que procesar
            if not lineas:
                messages.warning(request, "No se ingresó ninguna línea válida para registrar.")
                return render(request, "inventario/registrar_salida.html", {
                    "formset": formset,
                    "titulo": "Registrar Salida de Inventario",
                    "orden": orden_obj, 
                })
            
            for cd in lineas:
                insumo = cd.get("insumo")
                ubicacion = cd.get("ubicacion")
                cantidad = cd.get("cantidad") # Esto ya es Decimal
                fecha_salida = cd.get("fecha")
                lote = cd.get("insumo_lote") 
                observaciones = cd.get("observaciones", "")

                if lote is None or insumo is None or cantidad is None or ubicacion is None or fecha_salida is None or cantidad <= Decimal("0"):
                    messages.warning(request, f"Datos incompletos o cantidad inválida en línea para {insumo.nombre if insumo else 'un insumo'}. Se ignoró.")
                    continue 
                
                detalle_obj = None
                if orden_obj_post:
                    try:
                        # Buscamos el detalle de la orden para este insumo
                        detalle_obj = orden_obj_post.detalles.get(insumo=insumo)
                    except models.OrdenInsumoDetalle.DoesNotExist:
                        pass 

                # Aseguramos que la cantidad a salir no exceda el stock actual del lote
                cant = min(cantidad, lote.cantidad_actual or Decimal("0")) 
                
                models.Salida.objects.create(
                    insumo=insumo, insumo_lote=lote, ubicacion=ubicacion,
                    cantidad=cant, 
                    fecha_generada=fecha_salida, 
                    usuario=request.user,
                    tipo="USO_PRODUCCION", observaciones=observaciones,
                    orden=orden_obj_post,          
                    detalle=detalle_obj,      
                )
                
                # Actualizar stock del lote
                lote.cantidad_actual -= cant
                lote.save(update_fields=["cantidad_actual"])
                
                # Actualizar detalle de la orden
                if detalle_obj:
                    detalle_obj.cantidad_atendida += cant
                    detalle_obj.save(update_fields=["cantidad_atendida"])

            # Recalcular el estado de la orden al finalizar el loop de movimientos
            if orden_obj_post:
                orden_obj_post.recalc_estado()
                
            messages.success(request, "Salida(s) registrada(s) correctamente.")
            return redirect('inventario:listar_movimientos')

        messages.error(request, "Revisa los errores en el formulario antes de continuar.")

    else:
        # Evitar línea extra cuando se precarga desde insumo u orden
        if initial_data:
            # Usar extra=0 para que solo salgan las líneas precargadas
            DynamicFormSet = formset_factory(SalidaLineaForm, extra=0, can_delete=True) 
        else:
            # Usar extra=1 si no hay precarga, para dar un campo inicial al usuario
            DynamicFormSet = SalidaLineaFormSet
        formset = DynamicFormSet(initial=initial_data if initial_data else None)

    return render(request, "inventario/registrar_salida.html", {
        "formset": formset,
        "titulo": "Registrar Salida de Inventario",
        "orden": orden_obj, # Pasa la orden al contexto para la plantilla
    })


# --- REGISTRAR ENTRADA DESDE MOVIMIENTOS (CON LÍNEA AUTOMÁTICA) ---
@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def registrar_entrada_movimiento(request):
    """Registrar entrada CON línea automática (desde listar_movimientos)"""
    if not user_has_role(request.user, "Administrador", "Encargado"):
        messages.error(request, "No tienes permisos para registrar entradas.")
        return redirect('inventario:listar_movimientos')

    formset = None
    initial_data = []
    hay_precarga = False

    # --- LÓGICA DE PRE-CARGA (Insumo ID) ---
    insumo_id = request.GET.get("insumo_id")
    if insumo_id:
        try:
            insumo_obj = models.Insumo.objects.get(id=insumo_id)
            initial_data.append({
                'insumo': insumo_obj.id,
                'cantidad': 1,
            })
            hay_precarga = True
        except models.Insumo.DoesNotExist:
            messages.warning(request, "El insumo solicitado no es válido.")

    if request.method == "POST":
        # Usar formset con extra=1 (línea automática)
        formset = EntradaLineaFormSetMovimiento(request.POST)
        if formset.is_valid():
            lineas = [f.cleaned_data for f in formset if getattr(f, "cleaned_data", None) and not f.cleaned_data.get("DELETE")]
            
            # CRÍTICO: Bloquear el éxito si no hay líneas válidas que procesar
            if not lineas:
                messages.warning(request, "No se ingresó ninguna línea válida para registrar.")
                return render(request, "inventario/registrar_entrada.html", {
                    "formset": formset,
                    "titulo": "Registrar Entrada de Inventario",
                })
            
            for cd in lineas:
                insumo = cd.get("insumo")
                ubicacion = cd.get("ubicacion")
                cantidad = cd.get("cantidad")
                fecha = cd.get("fecha")
                proveedor = cd.get("proveedor")
                fecha_exp = cd.get("fecha_expiracion")
                
                if not all([insumo, ubicacion, cantidad, fecha, proveedor, fecha_exp]):
                    messages.warning(request, "Datos incompletos en una línea de entrada. Se ignoró esta línea.")
                    continue
                
                lote = models.InsumoLote.objects.create(
                    insumo=insumo,
                    bodega=ubicacion.bodega,
                    proveedor=proveedor,
                    fecha_ingreso=fecha,
                    fecha_expiracion=fecha_exp,
                    cantidad_inicial=cantidad,
                    cantidad_actual=Decimal("0.00"),
                    usuario=request.user,
                )

                models.Entrada.objects.create(
                    insumo=insumo, insumo_lote=lote, ubicacion=ubicacion,
                    cantidad=cantidad, fecha=fecha, usuario=request.user,
                    observaciones=cd.get("observaciones", ""),
                )
                
                lote.cantidad_actual += cantidad
                lote.save(update_fields=["cantidad_actual"])
                
                check_and_create_stock_alerts(insumo)

            messages.success(request, "✅ Entrada(s) registrada(s) correctamente.")
            return redirect('inventario:listar_movimientos')

        messages.error(request, "Revisa los errores en el formulario antes de continuar.")

    if formset is None:
        formset = EntradaLineaFormSetMovimiento(initial=initial_data if initial_data else None)

    return render(request, "inventario/registrar_entrada.html", {
        "formset": formset,
        "titulo": "Registrar Entrada de Inventario",
    })


# --- REGISTRAR SALIDA DESDE MOVIMIENTOS (CON LÍNEA AUTOMÁTICA) ---
@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def registrar_salida_movimiento(request):
    """Registrar salida CON línea automática (desde listar_movimientos)"""
    if not user_has_role(request.user, "Administrador", "Encargado"):
        messages.error(request, "No tienes permisos para registrar salidas.")
        return redirect('inventario:listar_movimientos')

    formset = None
    initial_data = []
    hay_precarga = False

    # --- LÓGICA DE PRE-CARGA (Orden ID) ---
    orden_id = request.GET.get("orden_id")
    orden_obj = None
    if orden_id:
        try:
            orden_obj = models.OrdenInsumo.objects.get(id=orden_id)
            hay_precarga = True
        except models.OrdenInsumo.DoesNotExist:
            messages.warning(request, "La orden solicitada no es válida.")

    if request.method == "POST":
        # Usar formset con extra=1 (línea automática)
        formset = SalidaLineaFormSetMovimiento(request.POST)
        if formset.is_valid():
            lineas = [f.cleaned_data for f in formset if getattr(f, "cleaned_data", None) and not f.cleaned_data.get("DELETE")]
            
            # CRÍTICO: Bloquear el éxito si no hay líneas válidas que procesar
            if not lineas:
                messages.warning(request, "No se ingresó ninguna línea válida para registrar.")
                return render(request, "inventario/registrar_salida.html", {
                    "formset": formset,
                    "titulo": "Registrar Salida de Inventario",
                    "orden": orden_obj, 
                })
            
            for cd in lineas:
                insumo = cd.get("insumo")
                ubicacion = cd.get("ubicacion")
                cantidad = cd.get("cantidad")
                fecha = cd.get("fecha")
                
                if not all([insumo, ubicacion, cantidad, fecha]):
                    messages.warning(request, "Datos incompletos en una línea de salida. Se ignoró esta línea.")
                    continue
                
                # Obtener lote disponible
                lote = models.InsumoLote.objects.filter(
                    insumo=insumo, bodega=ubicacion.bodega,
                    cantidad_actual__gt=0, is_active=True
                ).first()
                
                if not lote or lote.cantidad_actual < cantidad:
                    messages.warning(request, f"No hay stock disponible de {insumo.nombre} en {ubicacion.bodega.nombre}.")
                    continue
                
                models.Salida.objects.create(
                    insumo=insumo, insumo_lote=lote, ubicacion=ubicacion,
                    cantidad=cantidad, fecha=fecha, usuario=request.user,
                    observaciones=cd.get("observaciones", ""),
                )
                
                lote.cantidad_actual -= cantidad
                lote.save(update_fields=["cantidad_actual"])
                
                check_and_create_stock_alerts(insumo)

            messages.success(request, "✅ Salida(s) registrada(s) correctamente.")
            return redirect('inventario:listar_movimientos')

        messages.error(request, "Revisa los errores en el formulario antes de continuar.")

    if formset is None:
        formset = SalidaLineaFormSetMovimiento(initial=initial_data if initial_data else None)

    return render(request, "inventario/registrar_salida.html", {
        "formset": formset,
        "titulo": "Registrar Salida de Inventario",
        "orden": orden_obj,
    })
    
@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def registrar_entrada_orden(request, pk):
    """
    Vista dedicada a atender completamente una Orden de Insumo de tipo ENTRADA.
    Sigue la lógica de registrar_salida_orden para la precarga robusta.
    """
    from decimal import Decimal
    from django.forms import formset_factory 
    from django.db.models import F
    from datetime import date, timedelta
    from .forms import EntradaLineaForm, EntradaLineaFormSet # Aseguramos importación
    
    if not user_has_role(request.user, "Administrador", "Encargado"):
        messages.error(request, "No tienes permisos para registrar entradas de orden.")
        return redirect('inventario:listar_ordenes')

    orden_obj = get_object_or_404(models.OrdenInsumo, id=pk)

    # ⚠️ VALIDACIÓN CRÍTICA: Solo para órdenes de ENTRADA
    if orden_obj.tipo_orden != 'ENTRADA':
        messages.error(request, f"La orden #{orden_obj.id} es de tipo '{orden_obj.tipo_orden}'. Solo las órdenes de ENTRADA pueden atenderse aquí.")
        return redirect('inventario:listar_ordenes')

    # --- OBTENER DEFAULTS para campos obligatorios (Ubicación/Proveedor/Fecha) ---
    default_ubicacion = models.Ubicacion.objects.filter(is_active=True).first()
    default_proveedor = models.Proveedor.objects.filter(is_active=True).first()
    
    if not default_ubicacion or not default_proveedor:
         messages.error(request, "Error de Configuración: No hay ubicaciones o proveedores activos para precargar la entrada. Debe crear al menos uno.")
         return redirect('inventario:listar_ordenes')
    
    default_ubicacion_id = default_ubicacion.id
    default_proveedor_id = default_proveedor.id
    
    hoy = date.today()
    fecha_exp_default = hoy + timedelta(days=365) # Fecha de expiración por defecto
    
    initial_data = []

    # --- LÓGICA DE PRE-CARGA (GET) ---
    if request.method == "GET":
        # Precargar solo los detalles que no están cerrados
        for detalle in orden_obj.detalles.filter(cantidad_solicitada__gt=F('cantidad_atendida')): 
            
            # CHECK: Asegurar que el Insumo de la orden está ACTIVO
            insumo_activo = models.Insumo.objects.filter(id=detalle.insumo.id, is_active=True).exists()
            
            if not insumo_activo:
                messages.warning(request, f"El insumo '{detalle.insumo.nombre}' fue omitido: no está activo en el catálogo.")
                continue

            cantidad_restante = detalle.cantidad_solicitada - detalle.cantidad_atendida
            
            # 🔥 CORRECCIÓN FINAL: Convertir el Decimal a float (o str) para FormSet robusto
            cantidad_inicial = float(cantidad_restante)
            
            # CRÍTICO: Convertir TODOS los IDs a STRING para evitar fallos con múltiples líneas
            initial_data.append({
                'insumo': str(detalle.insumo.id),                      
                'cantidad': cantidad_inicial,                          # <-- Ahora es float
                'ubicacion': str(default_ubicacion_id),                
                'proveedor': str(default_proveedor_id),                
                'fecha': hoy.isoformat(),                              
                'fecha_expiracion': fecha_exp_default.isoformat(),     
            })
        
    if request.method == "POST":
        formset = EntradaLineaFormSet(request.POST) 
        orden_obj_post = orden_obj 

        if formset.is_valid():
            lineas = [f.cleaned_data for f in formset if getattr(f, "cleaned_data", None) and not f.cleaned_data.get("DELETE")]
            
            if not lineas:
                messages.warning(request, "No se ingresó ninguna línea válida para registrar.")
                # Si falla, redirigimos para reiniciar la precarga y mostrar el mensaje de advertencia.
                return redirect('inventario:registrar_entrada_orden', pk=pk)
            
            for cd in lineas:
                insumo = cd.get("insumo")
                ubicacion = cd.get("ubicacion")
                cantidad = cd.get("cantidad")
                fecha = cd.get("fecha")
                proveedor = cd.get("proveedor")
                fecha_exp = cd.get("fecha_expiracion")
                
                if not all([insumo, ubicacion, cantidad, fecha, proveedor, fecha_exp]):
                    messages.warning(request, "Datos incompletos en una línea de entrada. Se ignoró esta línea.")
                    continue
                
                detalle_obj = None
                try:
                    detalle_obj = orden_obj_post.detalles.get(insumo=insumo)
                    cantidad_max = detalle_obj.cantidad_solicitada - detalle_obj.cantidad_atendida
                    if cantidad > cantidad_max and cantidad_max > Decimal("0"):
                        cantidad = cantidad_max
                    elif cantidad_max <= Decimal("0"):
                        continue

                except models.OrdenInsumoDetalle.DoesNotExist:
                    messages.error(request, f"El insumo {insumo.nombre} no es parte de la Orden #{pk}. Abortando operación para evitar inconsistencias.")
                    raise Exception("Insumo no encontrado en la orden.")

                lote = models.InsumoLote.objects.create(
                    insumo=insumo,
                    bodega=ubicacion.bodega,
                    proveedor=proveedor,
                    fecha_ingreso=fecha,
                    fecha_expiracion=fecha_exp,
                    cantidad_inicial=cantidad,
                    cantidad_actual=Decimal("0.00"),
                    usuario=request.user,
                )

                models.Entrada.objects.create(
                    insumo=insumo, insumo_lote=lote, ubicacion=ubicacion,
                    cantidad=cantidad, fecha=fecha, usuario=request.user,
                    observaciones=cd.get("observaciones", ""),
                    orden=orden_obj_post,     
                    detalle=detalle_obj,      
                )
                
                lote.cantidad_actual += cantidad
                lote.save(update_fields=["cantidad_actual"])
                
                if detalle_obj:
                    detalle_obj.cantidad_atendida += cantidad
                    detalle_obj.save(update_fields=["cantidad_atendida"])
                
                models.check_and_create_stock_alerts(insumo)

            orden_obj_post.recalc_estado()
                
            messages.success(request, f"✅ Orden de Entrada #{pk} atendida correctamente. Movimiento(s) registrado(s).")
            return redirect('inventario:listar_ordenes')

        messages.error(request, "Revisa los errores en el formulario antes de continuar.")

    # --- LÓGICA RENDER (GET/Fallo POST) ---
    DynamicFormSet = formset_factory(EntradaLineaForm, extra=0, can_delete=True) 

    if request.method == "POST" and formset.errors:
        # Si el POST falla, el formset ya contiene los errores
        pass 
    else:
        # Si es GET, inicializamos desde initial_data (que contiene las líneas precargadas)
        formset = DynamicFormSet(initial=initial_data if initial_data else None)

    return render(request, "inventario/registrar_entrada.html", {
        "formset": formset,
        "titulo": f"Atender Orden de Entrada #{pk}",
        "orden": orden_obj, 
    })
    
@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def registrar_salida_orden(request, pk):
    if not user_has_role(request.user, "Administrador", "Encargado"):
        messages.error(request, "No tienes permisos para registrar salidas de orden.")
        return redirect('inventario:listar_ordenes')

    initial_data = []
    orden_obj = get_object_or_404(models.OrdenInsumo, id=pk)

    # ⚠️ VALIDACIÓN CRÍTICA: Solo para órdenes de SALIDA
    if orden_obj.tipo_orden != 'SALIDA':
        messages.error(request, f"La orden #{orden_obj.id} es de tipo '{orden_obj.tipo_orden}'. Solo las órdenes de SALIDA pueden atenderse aquí.")
        return redirect('inventario:listar_ordenes')

    # --- LÓGICA DE PRE-CARGA (GET) ---
    if request.method == "GET":
        # Filtrar solo los detalles pendientes (solicitada > atendida)
        detalles_pendientes = orden_obj.detalles.filter(
            cantidad_solicitada__gt=F('cantidad_atendida')
        ).select_related('insumo')

        for detalle in detalles_pendientes:
            cantidad_restante = detalle.cantidad_solicitada - detalle.cantidad_atendida
            
            # Usamos la cantidad restante (Decimal) directamente para el initial data
            cantidad_inicial = cantidad_restante

            # Añadir solo si la cantidad inicial es positiva
            if cantidad_inicial > Decimal("0.00"):
                initial_data.append({
                    'insumo': detalle.insumo.id, 
                    'cantidad': cantidad_inicial, 
                })
        
    if request.method == "POST":
        # Usamos el formset con los datos POST
        formset = SalidaLineaFormSet(request.POST) 
        orden_obj_post = orden_obj # La orden ya fue verificada

        if formset.is_valid():
            lineas = [f.cleaned_data for f in formset if getattr(f, "cleaned_data", None) and not f.cleaned_data.get("DELETE")]
            
            if not lineas:
                messages.warning(request, "No se ingresó ninguna línea válida para registrar.")
                # Si el POST falla sin líneas válidas, volvemos a mostrar el formulario con extra=0
                DynamicFormSet = formset_factory(SalidaLineaForm, extra=0, can_delete=True)
                formset = DynamicFormSet(initial=initial_data if initial_data else None)
                return render(request, "inventario/registrar_salida.html", {
                    "formset": formset,
                    "titulo": f"Atender Orden de Salida #{pk}",
                    "orden": orden_obj, 
                })
            
            for cd in lineas:
                insumo = cd.get("insumo")
                ubicacion = cd.get("ubicacion")
                cantidad = cd.get("cantidad")
                fecha_salida = cd.get("fecha")
                lote = cd.get("insumo_lote") 
                observaciones = cd.get("observaciones", "")

                if lote is None or insumo is None or cantidad is None or ubicacion is None or fecha_salida is None or cantidad <= Decimal("0"):
                    messages.warning(request, f"Datos incompletos o cantidad inválida en línea para {insumo.nombre if insumo else 'un insumo'}. Se ignoró.")
                    continue 
                
                # --- ENCONTRAR DETALLE Y VINCULAR ---
                detalle_obj = None
                try:
                    # Buscamos el detalle de la orden para este insumo
                    detalle_obj = orden_obj_post.detalles.get(insumo=insumo)
                except models.OrdenInsumoDetalle.DoesNotExist:
                    messages.error(request, f"El insumo {insumo.nombre} no es parte de la Orden #{pk}. Abortando operación para evitar inconsistencias.")
                    # Si el insumo no estaba en la orden original, elevamos una excepción para abortar la transacción atómica
                    raise Exception("Insumo no encontrado en la orden.")
                
                # Aseguramos que la cantidad a salir no exceda el stock actual del lote
                cant = min(cantidad, lote.cantidad_actual or Decimal("0")) 
                
                models.Salida.objects.create(
                    insumo=insumo, insumo_lote=lote, ubicacion=ubicacion,
                    cantidad=cant, 
                    fecha_generada=fecha_salida, 
                    usuario=request.user,
                    tipo="USO_PRODUCCION", observaciones=observaciones,
                    orden=orden_obj_post,          
                    detalle=detalle_obj,      
                )
                
                # Actualizar stock del lote
                lote.cantidad_actual -= cant
                lote.save(update_fields=["cantidad_actual"])
                
                # Actualizar detalle de la orden
                if detalle_obj:
                    detalle_obj.cantidad_atendida += cant
                    detalle_obj.save(update_fields=["cantidad_atendida"])
                
                # Recalcular alertas
                check_and_create_stock_alerts(insumo)


            # Recalcular el estado de la orden al finalizar el loop de movimientos
            orden_obj_post.recalc_estado()
                
            messages.success(request, f"✅ Orden de Salida #{pk} atendida correctamente. Movimiento(s) registrado(s).")
            return redirect('inventario:listar_ordenes')

        messages.error(request, "Revisa los errores en el formulario antes de continuar.")

    # --- LÓGICA RENDER (GET/Fallo POST) ---
    # Usamos extra=0 en ambos casos para evitar que se añadan líneas no relacionadas con la orden.
    DynamicFormSet = formset_factory(SalidaLineaForm, extra=0, can_delete=True) 

    formset = DynamicFormSet(initial=initial_data if initial_data else None)

    return render(request, "inventario/registrar_salida.html", {
        "formset": formset,
        "titulo": f"Atender Orden de Salida #{pk}",
        "orden": orden_obj, # Pasa la orden al contexto
    })
    
    
@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def editar_entrada(request, pk):
    # 1. Recuperar la entrada
    entrada = get_object_or_404(models.Entrada, pk=pk)

    # --- 2. INICIALIZACIÓN DE OBJETOS HISTÓRICOS Y MANEJO DE ERROR ---
    
    try:
        # Intentamos obtener el Insumo y el Lote directamente por PK
        historic_insumo = models.Insumo.objects.get(pk=entrada.insumo_id)
        historic_lote = models.InsumoLote.objects.get(pk=entrada.insumo_lote_id)
    except (models.Insumo.DoesNotExist, models.InsumoLote.DoesNotExist):
        # Si el objeto maestro no existe, lo notificamos y terminamos la edición.
        messages.error(request, f"No se puede editar el movimiento #{pk}: El insumo o lote asociado ya no existe en el sistema.")
        return redirect("inventario:listar_movimientos")

    # 3. Inicializar el formulario
    form = EntradaForm(request.POST or None, instance=entrada)
    
    # 4. Inyectar dinámicamente los objetos históricos al QuerySet del Formulario
    # Esto evita el error "Insumo matching query does not exist." durante la inicialización.
    
    # A) Ajustar el queryset del campo 'insumo'
    insumo_field = form.fields['insumo']
    # Si el insumo histórico no está en el queryset actual (ej. por filtro is_active), lo añadimos usando OR.
    if not insumo_field.queryset.filter(pk=historic_insumo.pk).exists():
        insumo_field.queryset = insumo_field.queryset | models.Insumo.objects.filter(pk=historic_insumo.pk)

    # B) Ajustar el queryset del campo 'insumo_lote'
    lote_field = form.fields['insumo_lote']
    if not lote_field.queryset.filter(pk=historic_lote.pk).exists():
        lote_field.queryset = lote_field.queryset | models.InsumoLote.objects.filter(pk=historic_lote.pk)


    # --- 5. RESTO DE LA LÓGICA DE POST ---
    old_qty = Decimal(entrada.cantidad)
    lote = entrada.insumo_lote

    if request.method == "POST" and form.is_valid():
        nueva = Decimal(form.cleaned_data["cantidad"])
        delta = nueva - old_qty
        
        # Validación de stock futuro
        if lote.cantidad_actual + delta < 0:
            messages.error(request, f"Error al editar: La nueva cantidad excedería el stock del lote, resultando en stock negativo ({lote.cantidad_actual + delta:.2f}).")
            return render(request, "inventario/editar_movimiento.html", {"form": form, "titulo": "Editar Entrada"})

        lote.cantidad_actual += delta
        lote.save(update_fields=["cantidad_actual"])
        form.save()

        if entrada.detalle_id:
            det = entrada.detalle
            det.cantidad_atendida += delta
            if det.cantidad_atendida < 0:
                det.cantidad_atendida = 0
            det.save(update_fields=["cantidad_atendida"])
            entrada.orden.recalc_estado()

        messages.success(request, "Entrada modificada correctamente.")
        return redirect("inventario:listar_movimientos")

    return render(
    request,
    "inventario/editar_movimiento.html",
    {
        "form": form,
        "titulo": "Editar Entrada",
        "cancel_url": reverse("inventario:listar_movimientos"),
        "campo_fecha": form["fecha"],                 
        "errores_fecha": form["fecha"].errors,       
    },
)


@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def eliminar_entrada(request, pk):
    entrada = get_object_or_404(Entrada, pk=pk)
    
    # Se añade validación para evitar stock negativo al eliminar la entrada
    lote = entrada.insumo_lote
    if lote.cantidad_actual - entrada.cantidad < 0:
        messages.error(request, f"Error: No se puede eliminar esta entrada ya que reduciría el stock del lote a un valor negativo ({lote.cantidad_actual - entrada.cantidad:.2f}).")
        return redirect("inventario:listar_movimientos") # Redirige con error

    if request.method == "POST":
        lote.cantidad_actual -= entrada.cantidad
        lote.save(update_fields=["cantidad_actual"])
        if entrada.detalle_id:
            det = entrada.detalle
            det.cantidad_atendida -= entrada.cantidad
            if det.cantidad_atendida < 0:
                det.cantidad_atendida = 0
            det.save(update_fields=["cantidad_atendida"])
            entrada.orden.recalc_estado()
        entrada.delete()
        messages.success(request, "Entrada eliminada correctamente.")
        return redirect("inventario:listar_movimientos")

    return render(request, "inventario/confirmar_eliminar.html", {"obj": entrada, "tipo": "entrada"})


@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def editar_salida(request, pk):
    salida = get_object_or_404(Salida, pk=pk)
    old_qty = Decimal(salida.cantidad)
    lote = salida.insumo_lote

    # REMOVIDA: from .forms import SalidaForm (ahora importada globalmente)
    form = SalidaForm(request.POST or None, instance=salida)
    if request.method == "POST" and form.is_valid():
        nueva = Decimal(form.cleaned_data["cantidad"])
        delta = nueva - old_qty
        
        # Validación de stock: la diferencia entre la cantidad actual en lote y el delta
        # no debe ser menor a cero. Si delta es positivo, es una mayor salida, reduce stock.
        if lote.cantidad_actual - delta < 0:
            messages.error(request, f"Error al editar: La nueva salida excede el stock actual del lote, resultando en stock negativo ({lote.cantidad_actual - delta:.2f}).")
            return render(request, "inventario/editar_movimiento.html", {"form": form, "titulo": "Editar Salida"})

        lote.cantidad_actual -= delta
        lote.save(update_fields=["cantidad_actual"])
        form.save()
        messages.success(request, "Salida modificada correctamente.")
        return redirect("inventario:listar_movimientos")

    return render(
    request,
    "inventario/editar_movimiento.html",
    {
        "form": form,
        "titulo": "Editar Salida",
        "cancel_url": reverse("inventario:listar_movimientos"),
        "campo_fecha": form["fecha_generada"],        # ✅ BoundField seguro
        "errores_fecha": form["fecha_generada"].errors,
    },
)


@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def eliminar_salida(request, pk):
    salida = get_object_or_404(Salida, pk=pk)
    if request.method == "POST":
        lote = salida.insumo_lote
        # Al eliminar una salida, se revierte el stock
        lote.cantidad_actual += salida.cantidad
        lote.save(update_fields=["cantidad_actual"])
        salida.delete()
        messages.success(request, "Salida eliminada correctamente.")
        return redirect("inventario:listar_movimientos")

    return render(request, "inventario/confirmar_eliminar.html", {"obj": salida, "tipo": "salida"})

# --- LISTAR MOVIMIENTOS (ENTRADAS Y SALIDAS) ---
@login_required
@perfil_required(allow=("administrador", "Encargado")) 
def listar_movimientos(request):
    titulo = "Movimientos de Inventario"
    q = (request.GET.get("q") or "").strip()

    entradas_qs = Entrada.objects.select_related("insumo", "insumo_lote", "ubicacion").order_by("-fecha")
    salidas_qs  = Salida.objects.select_related("insumo", "insumo_lote", "ubicacion").order_by("-fecha_generada")

    if q:
        entradas_qs = entradas_qs.filter(
            Q(insumo__nombre__icontains=q) | Q(ubicacion__nombre__icontains=q) | Q(observaciones__icontains=q)
        )
        salidas_qs = salidas_qs.filter(
            Q(insumo__nombre__icontains=q) | Q(ubicacion__nombre__icontains=q)
        )

    p_e = Paginator(entradas_qs, 20)
    p_s = Paginator(salidas_qs, 20)
    page_e = request.GET.get("page_e")
    page_s = request.GET.get("page_s")
    entradas = p_e.get_page(page_e)
    salidas  = p_s.get_page(page_s)

    can_manage = request.user.is_superuser or user_has_role(request.user, "administrador", "encargado")

    return render(request, "inventario/listar_movimientos.html", {
        "titulo": titulo,
        "entradas": entradas,
        "salidas": salidas,
        "can_manage": can_manage,
        "q": q,
    })



# --- Bodegas
@login_required
@perfil_required(
    allow=("administrador", "Encargado"))
def listar_bodegas(request):
    qs = Bodega.objects.filter(is_active=True)

    allowed_sort = {"nombre", "direccion"}
    sort = request.GET.get("sort")
    if sort not in allowed_sort:
        sort = request.session.get("sort_bodegas", "nombre")
    if sort not in allowed_sort:
        sort = "nombre"
    request.session["sort_bodegas"] = sort

    # NUEVO: dirección de orden
    order = request.GET.get("order")
    if order not in ("asc", "desc"):
        order = request.session.get("order_bodegas", "asc")
    if order not in ("asc", "desc"):
        order = "asc"
    request.session["order_bodegas"] = order

    sort_map = {
        "nombre": "nombre",
        "direccion": "direccion",
    }

    read_only = not (request.user.is_superuser or user_has_role(request.user, "Administrador", "Encargado"))

    return list_with_filters(
        request,
        qs,
        search_fields=["nombre", "direccion"],
        order_field=sort_map[sort],
        session_prefix="bodegas",
        context_key="bodegas",
        full_template="inventario/listar_bodegas.html",
        partial_template="inventario/partials/bodegas_results.html",
        default_per_page=10,
        default_order="asc",   # GET 'order' la sobreescribe
        tie_break="id",
        extra_context={
            "titulo": "Bodegas",
            "sort": sort,
            "order": order,     # <-- importante
            "read_only": read_only,
        },
    )

class OrdenForm(ModelForm):
    class Meta:
        model = OrdenInsumo
        fields = ("estado",)  # el usuario lo fijamos con request.user

DetalleFormSet = inlineformset_factory(
    OrdenInsumo,
    OrdenInsumoDetalle,
    fields=("insumo", "cantidad_solicitada"),
    extra=1,
    can_delete=True
)

@login_required
@perfil_required(allow=("administrador","encargado"))
def crear_bodega(request):
    form = BodegaForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "✅ Bodega creada correctamente.")
        return redirect("inventario:listar_bodegas")
    return render(request, "inventario/bodega_form.html", {"form": form, "titulo": "Nueva Bodega"})

@login_required
@perfil_required(allow=("administrador","encargado"))
def editar_bodega(request, pk):
    obj = get_object_or_404(Bodega, pk=pk, is_active=True)
    form = BodegaForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "📝 Bodega actualizada.")
        return redirect("inventario:listar_bodegas")
    return render(request, "inventario/bodega_form.html", {"form": form, "titulo": f"Editar: {obj.nombre}"})

@login_required
@perfil_required(allow=("administrador",))
def eliminar_bodega(request, pk):
    bodega = get_object_or_404(Bodega, pk=pk)

    # Regla: si tiene lotes activos asociados, no se puede eliminar
    tiene_lotes = InsumoLote.objects.filter(bodega=bodega, is_active=True).exists()

    if request.method == "POST":
        if tiene_lotes:
            messages.error(request, "No se puede eliminar la bodega: tiene lotes activos asociados.")
            return redirect("inventario:listar_bodegas")
        bodega.is_active = False
        bodega.save(update_fields=["is_active"])
        messages.success(request, f"Bodega '{bodega.nombre}' eliminada.")
        return redirect("inventario:listar_bodegas")

    return render(
        request,
        "confirmar_eliminar.html",
        {"titulo": f"Eliminar Bodega «{bodega.nombre}»", "cancel_url": reverse("inventario:listar_bodegas")}
    )


# --- LISTAR ÓRDENES ---
ESTADOS_ORDEN = ("PENDIENTE", "EN_CURSO", "CERRADA", "CANCELADA")

@login_required
@perfil_required(allow=("administrador", "Encargado"), readonly_for=("Bodeguero",))
def listar_ordenes(request):
    """Lista las órdenes y permite actualizar su estado manualmente."""
    
    # MODIFICADO: Se utiliza prefetch_related para optimizar la carga de detalles
    qs = (
        OrdenInsumo.objects.all()
        .select_related("usuario")
        .prefetch_related(
            Prefetch(
                "detalles",
                queryset=OrdenInsumoDetalle.objects.select_related(
                    "insumo__unidad_medida"
                )
            )
        )
    )

    # --- Búsqueda (sin cambios)
    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(Q(usuario__name__icontains=q) | Q(usuario__email__icontains=q))

    # --- Filtro por estado (sin cambios)
    estado_f = request.GET.get("estado")
    if estado_f in ESTADOS_ORDEN:
        qs = qs.filter(estado=estado_f)

    # --- Orden (sin cambios)
    allowed_sort = {"id", "fecha", "estado", "usuario", "creado", "actualizado"}
    sort = request.GET.get("sort")
    if sort not in allowed_sort:
        sort = "fecha"

    sort_map = {
        "id": "id", "fecha": "fecha", "estado": "estado", "usuario": "usuario__name",
        "creado": "created_at", "actualizado": "updated_at",
    }

    # CORRECCIÓN CLAVE: El botón 'Nueva orden' se muestra si el usuario es Admin, Encargado, o Bodeguero.
    read_only = not (
        request.user.is_superuser
        or user_has_role(request.user, "Administrador", "Encargado", "Bodeguero") # <-- ¡BODEGUERO AÑADIDO AQUÍ!
    )

    return list_with_filters(
        request,
        qs,
        search_fields=[],
        order_field=sort_map[sort],
        session_prefix="ordenes",
        context_key="ordenes",
        full_template="inventario/listar_ordenes.html",
        partial_template="inventario/partials/ordenes_results.html",
        default_per_page=10,
        default_order="desc",
        tie_break="id",
        extra_context={
            "titulo": "Órdenes de Insumo",
            "q": q,
            "estado": estado_f,
            "sort": sort,
            "ESTADOS_ORDEN": ESTADOS_ORDEN,
            "ESTADO_ORDEN_CHOICES": ESTADO_ORDEN_CHOICES,
            "read_only": read_only, # <--- Se pasa el valor actualizado
        },
    )

@login_required
@perfil_required(allow=("administrador", "Encargado"))
def orden_cambiar_estado(request, pk):
    orden = get_object_or_404(OrdenInsumo, pk=pk)
    nuevo = request.POST.get("estado")
    if nuevo not in ESTADOS_ORDEN:
        return HttpResponseBadRequest("Estado inválido")

    if orden.estado != nuevo:
        orden.estado = nuevo
        orden.save(update_fields=["estado", "updated_at"])
        messages.success(request, f"Estado de la orden #{orden.id} actualizado a {nuevo}.")

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"ok": True})
    return redirect("inventario:listar_ordenes")

@login_required
@perfil_required(allow=("administrador", "Encargado", "Bodeguero")) # <-- El rol 'Bodeguero' está incluido aquí
@transaction.atomic
def crear_orden(request):
    """
    Crea una nueva orden con mínimo 1 ítem válido.
    """
    orden = OrdenInsumo()  # Aún no guardada
    
    # NUEVO: Formulario principal para el tipo de orden
    form = OrdenInsumoForm(request.POST or None)
    
    formset = OrdenInsumoDetalleCreateFormSet(request.POST or None, instance=orden)

    if request.method == "POST":

        if form.is_valid() and formset.is_valid(): # <--- Validar ambos formularios
            
            # Guardar el formulario principal (Obtiene el tipo_orden)
            orden = form.save(commit=False)
            orden.usuario = request.user
            orden.save()
            
            # Guardar los detalles con la instancia de orden guardada
            formset.instance = orden
            formset.save()

            messages.success(request, "Orden creada correctamente.")
            return redirect("inventario:listar_ordenes")

        messages.error(request, "Corrige los errores antes de continuar.")

    return render(
        request,
        "inventario/crear_orden.html",
        {
            "titulo": "Crear orden",
            "form": form,      # <--- Pasar el formulario principal
            "formset": formset,
        }
    )


@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def editar_orden(request, pk):
    """
    Edita una orden existente. Permite eliminar detalles y agregar nuevos.
    """
    orden = get_object_or_404(OrdenInsumo, pk=pk)
    
    # NUEVO: Formulario principal (el campo tipo_orden se deshabilita automáticamente)
    form = OrdenInsumoForm(request.POST or None, instance=orden)
    
    formset = OrdenInsumoDetalleEditFormSet(request.POST or None, instance=orden)

    if request.method == "POST":

        # Validación: si la orden está CERRADA o CANCELADA, se prohíbe la edición de detalles
        if orden.estado in ["CERRADA", "CANCELADA"]:
            messages.error(request, f"La orden #{orden.pk} está {orden.estado} y no puede ser editada.")
            return redirect("inventario:listar_ordenes")


        if form.is_valid() and formset.is_valid():
            
            form.save() # Guarda el estado y actualiza updated_at

            formset.save()

            messages.success(request, f"Orden #{orden.pk} actualizada correctamente.")
            return redirect("inventario:listar_ordenes")

        messages.error(request, "Revisa los errores en los ítems de la orden.")

    return render(
        request,
        "inventario/crear_orden.html",
        {
            "titulo": f"Editar orden #{orden.pk} ({orden.tipo_orden})", # MODIFICADO título
            "form": form,      # <--- Pasar el formulario principal
            "formset": formset,
        }
    )


@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def eliminar_orden(request, pk):
    orden = get_object_or_404(OrdenInsumo, pk=pk)
    if not user_has_role(request.user, "administrador", "Encargado"):
        messages.error(request, "No tienes permisos para eliminar órdenes.")
        return redirect("inventario:listar_ordenes")

    if request.method == "POST":
        orden.delete()
        messages.success(request, "Orden eliminada.")
        return redirect("inventario:listar_ordenes")

    # puedes reutilizar una plantilla genérica de confirmación
    return render(request, "inventario/confirmar_eliminar.html", {
        "titulo": f"Eliminar Orden #{orden.id}",
        "obj": orden,
        "tipo": "orden",
    })

# --- REPORTES ---
@login_required
@perfil_required(allow=("administrador", "Encargado"))
def reporte_disponibilidad(request):
    """
    Reporte de disponibilidad de insumos con:
    - Filtro por insumos (nombre) via checkboxes.
    - Flags de columnas a mostrar.
    - Lotes indentados por insumo (en HTML).
    - Export: CSV/XLSX/PDF (respetan filtro de insumos).
    - Soporte AJAX para actualizar solo la tabla.
    """
    hoy = date.today()

    # -------- flags de columnas (HTML) --------
    # por defecto mostramos: stock_total, precio_unitario, prox_vencimiento, lotes
    def _b(name, default=True):
        """
        Lee tanto 'show_*' (nuevo) como 'col_*' (legado) y devuelve bool.
        """
        raw = request.GET.get(name, None)
        if raw is None and name.startswith("show_"):
            raw = request.GET.get(name.replace("show_", "col_"), None)  # compat con col_*
        if raw is None:
            return default
        return str(raw).lower() in ("1", "true", "on", "yes")

    show_precio_unitario = _b("show_precio_unitario", True)
    show_stock_total     = _b("show_stock_total", True)
    show_prox_venc       = _b("show_prox_venc", True)
    show_lotes           = _b("show_lotes", True)
    show_categorias      = _b("show_categorias", False)
    show_precio_acum     = _b("show_precio_acum", False)

    # -------- selección por insumo (nombres) --------
    selected_insumos = request.GET.getlist("insumo")   # lista de nombres exactos

    # Base queryset de insumos (activos + categoría activa)
    insumos_base = (
        Insumo.objects.filter(is_active=True, categoria__is_active=True)
        .select_related("categoria")
    )

    # Si vienen seleccionados, filtramos por nombre (case-insensitive)
    if selected_insumos:
        insumos_base = insumos_base.filter(nombre__in=selected_insumos)

    # Prefetch de lotes activos para cada insumo (ordenados por fecha de expiración)
    lotes_qs = (
        InsumoLote.objects
        .filter(is_active=True)
        .select_related("bodega")
        .order_by("fecha_expiracion", "id")
    )

    # Anotaciones de stock total, #lotes con stock, prox venc
    insumos_qs = (
        insumos_base
        .annotate(
            stock_total=Coalesce(
                Sum(
                    "lotes__cantidad_actual",
                    filter=Q(lotes__is_active=True),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
                0,
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            lotes_con_stock=Count(
                "lotes",
                filter=Q(lotes__is_active=True, lotes__cantidad_actual__gt=0),
                distinct=True,
            ),
            prox_vencimiento=Min(
                "lotes__fecha_expiracion",
                filter=Q(lotes__is_active=True, lotes__cantidad_actual__gt=0),
            ),
        )
        .prefetch_related(Prefetch("lotes", queryset=lotes_qs, to_attr="lotes_vis"))
        .order_by("categoria__nombre", "nombre")
    )

    # Dataset para checkboxes (todos los nombres disponibles, ordenados)
    all_insumo_names = list(
        Insumo.objects.filter(is_active=True, categoria__is_active=True)
        .order_by("nombre")
        .values_list("nombre", flat=True)
    )

    # Agrupado por categoría (para HTML)
    categorias = []
    cat_actual = None
    buffer = []
    for i in insumos_qs:
        if not cat_actual or i.categoria_id != cat_actual.id:
            if buffer:
                categorias.append({"categoria": cat_actual, "insumos": buffer})
                buffer = []
            cat_actual = i.categoria
        # cálculo auxiliar para HTML
        i.precio_acumulado = (i.stock_total or 0) * (i.precio_unitario or 0)
        buffer.append(i)
    if buffer:
        categorias.append({"categoria": cat_actual, "insumos": buffer})

    total_stock = sum((i.stock_total or 0) for i in insumos_qs)
    total_valor = sum(
        ((i.stock_total or 0) * (i.precio_unitario or 0))
        for i in insumos_qs
    )

    fmt = (request.GET.get("format") or "").lower()

    # ------- EXPORTS (respetan filtro de insumos, mantienen columnas clásicas) -------
    if fmt == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="reporte_disponibilidad_{hoy.isoformat()}.csv"'
        )
        writer = csv.writer(response)
        # misma estructura de columnas que PDF (más categoría)
        writer.writerow(
            [
                "Categoría",
                "Insumo",
                "Unidad",
                "Precio Unitario",
                "Stock Total",
                "Lotes con Stock",
                "Próx. Vencimiento",
            ]
        )
        for bloque in categorias:
            for i in bloque["insumos"]:
                writer.writerow(
                    [
                        bloque["categoria"].nombre if bloque["categoria"] else "",
                        i.nombre,
                        str(i.unidad_medida), # Asegurar str() para CSV
                        f"{i.precio_unitario}",
                        f"{i.stock_total:.2f}",
                        i.lotes_con_stock,
                        i.prox_vencimiento.isoformat()
                        if i.prox_vencimiento
                        else "—",
                    ]
                )
        writer.writerow([])
        writer.writerow(
            ["", "", "", "TOTALES", f"{Decimal(total_stock):.2f}", "", ""]
        )
        writer.writerow(
            ["", "", "", "VALOR TOTAL", f"{Decimal(total_valor):.2f}", "", ""]
        )
        return response

    # ---------- XLSX ----------
    if fmt in ("xlsx", "excel"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Disponibilidad"

        # Define Fills and Fonts
        HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type="solid") # Light Gray
        ROW_ODD_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type="solid") # Very Light Gray
        ROW_EVEN_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid") # White
        TOTAL_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid") # Light Green (for totals)
        BOLD_FONT = Font(bold=True)
        
        # --- TÍTULO ---
        title = f"Reporte de Disponibilidad de Insumos - {hoy.isoformat()}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=14, bold=True, color='000000') # Black color
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # --- CABECERAS ---
        headers = [
            "Categoría",
            "Insumo",
            "Unidad",
            "Precio Unitario",
            "Stock Total",
            "Lotes con Stock",
            "Próx. Vencimiento",
        ]
        ws.append(headers)

        header_row = ws[2]
        for cell in header_row:
            cell.font = BOLD_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = HEADER_FILL

        # --- DATOS Y ALTERNANCIA DE COLOR ---
        row_num = 3
        for bloque in categorias:
            for i in bloque["insumos"]:
                
                # Determine row style based on row_num parity
                fill = ROW_ODD_FILL if (row_num % 2 == 1) else ROW_EVEN_FILL
                
                ws.append(
                    [
                        bloque["categoria"].nombre,
                        i.nombre,
                        str(i.unidad_medida), # CORREGIDO: Forzar a string
                        float(i.precio_unitario or 0),
                        float(i.stock_total or 0),
                        int(i.lotes_con_stock or 0),
                        (
                            i.prox_vencimiento.isoformat()
                            if i.prox_vencimiento
                            else "—"
                        ),
                    ]
                )
                
                # Apply styles to the newly added row
                for col_idx in range(1, 8):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.fill = fill
                    
                    # Align text columns left
                    if col_idx in [1, 2, 3]:
                        cell.alignment = Alignment(horizontal="left")
                    # Align numbers right/center
                    elif col_idx in [4, 5]:
                        cell.alignment = Alignment(horizontal="right")
                    elif col_idx in [6, 7]:
                        cell.alignment = Alignment(horizontal="center")
                
                row_num += 1

        # --- TOTALES ---
        # 1. Add empty row for spacing
        ws.append([]) 
        total_stock_row_idx = ws.max_row + 1

        # 2. Total Stock Row
        ws.append(["", "", "", "TOTAL STOCK", float(total_stock), "", ""])
        total_stock_row_idx = ws.max_row
        
        # 3. Total Value Row
        ws.append(["", "", "", "VALOR TOTAL", float(total_valor), "", ""])
        total_value_row_idx = ws.max_row

        # Apply total styles (bold font and light green fill)
        for row_idx in [total_stock_row_idx, total_value_row_idx]:
            for col_idx in range(4, 6): # Columns D and E
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = BOLD_FONT
                cell.fill = TOTAL_FILL
                cell.alignment = Alignment(horizontal="right")
            
            # Apply bold to the label column (C)
            label_cell = ws.cell(row=row_idx, column=4)
            label_cell.font = BOLD_FONT
            label_cell.fill = TOTAL_FILL
            label_cell.alignment = Alignment(horizontal="left")


        # --- FORMATOS NUMÉRICOS (Aplicar a toda la columna D y E, incluyendo totales) ---
        
        # Curreny format for column D (Precio Unitario)
        for row in ws.iter_rows(min_row=3, min_col=4, max_col=4, max_row=ws.max_row):
            for cell in row:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE # e.g., $1,000.00
                
        # Decimal format for column E (Stock Total)
        for row in ws.iter_rows(min_row=3, min_col=5, max_col=5, max_row=ws.max_row):
            for cell in row:
                cell.number_format = "0.00"

        # --- ANCHO DE COLUMNAS ---
        col_widths = [18, 30, 15, 18, 16, 18, 18] # Adjusted based on content
        for idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
            
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        response = HttpResponse(
            bio.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename="reporte_disponibilidad_{hoy.isoformat()}.xlsx"'
        return response

    # ---------- PDF ----------
    if fmt == "pdf":
        bio = BytesIO()
        doc = SimpleDocTemplate(
            bio,
            pagesize=landscape(A4),
            leftMargin=20,
            rightMargin=20,
            topMargin=20,
            bottomMargin=20,
            title="Reporte de Disponibilidad de Insumos",
        )
        styles = getSampleStyleSheet()
        story = []

        story.append(
            Paragraph(
                f"Reporte de Disponibilidad de Insumos - {hoy.isoformat()}",
                styles["Title"],
            )
        )
        story.append(Spacer(1, 8))

        table_head = [
            "Insumo",
            "Unidad",
            "Precio Unitario",
            "Stock Total",
            "Lotes con Stock",
            "Próx. Vencimiento",
        ]

        for bloque in categorias:
            story.append(
                Paragraph(
                    f"Categoría: {bloque['categoria'].nombre}",
                    styles["Heading3"],
                )
            )
            data = [table_head]
            for i in bloque["insumos"]:
                data.append(
                    [
                        i.nombre,
                        str(i.unidad_medida), # Asegurar str() para PDF
                        f"{(i.precio_unitario or 0):,.0f}",
                        f"{(i.stock_total or 0):,.2f}",
                        int(i.lotes_con_stock or 0),
                        (
                            i.prox_vencimiento.strftime("%Y-%m-%d")
                            if i.prox_vencimiento
                            else "—"
                        ),
                    ]
                )
            t = Table(
                data, colWidths=[140, 60, 90, 90, 90, 110]
            )
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                        ("ALIGN", (2, 1), (-2, -1), "RIGHT"),
                        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.whitesmoke, colors.white],
                        ),
                    ]
                )
            )
            story.append(t)
            story.append(Spacer(1, 10))

        # Totales
        story.append(Spacer(1, 6))
        story.append(
            Paragraph(
                f"<b>Stock total:</b> {Decimal(total_stock):,.2f}",
                styles["Normal"],
            )
        )
        story.append(
            Paragraph(
                f"<b>Precio total:</b> {Decimal(total_valor):,.0f}",
                styles["Normal"],
            )
        )

        doc.build(story)
        pdf_value = bio.getvalue()
        bio.close()
        response = HttpResponse(pdf_value, content_type="application/pdf")
        response[
            "Content-Disposition"
        ] = f'attachment; filename=\"reporte_disponibilidad_{hoy.isoformat()}.pdf\"'
        return response

    # ---------- HTML (por defecto) ----------
    colspan_lotes = (
        2
        + int(show_categorias)
        + int(show_precio_unitario)
        + int(show_stock_total)
        + int(show_prox_venc)
        + int(show_precio_acum)
    )

    context = {
        "categorias": categorias,
        "total_stock": total_stock,
        "total_valor": total_valor,
        "hoy": hoy,
        "titulo": "Reporte de Disponibilidad de Insumos",
        # flags UI
        "show_precio_unitario": show_precio_unitario,
        "show_stock_total": show_stock_total,
        "show_prox_venc": show_prox_venc,
        "show_lotes": show_lotes,
        "show_categorias": show_categorias,
        "show_precio_acum": show_precio_acum,
        # selección de insumos
        "all_insumo_names": all_insumo_names,
        "selected_insumos": set(selected_insumos),
        "colspan_lotes": colspan_lotes,
        "request": request,
    }

    # Si es AJAX, devolvemos solo el fragmento HTML de la tabla
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        html = render_to_string(
            "inventario/partials/reporte_disponibilidad_results.html",
            context,
            request=request,
        )
        return JsonResponse({"ok": True, "html": html})

    return render(request, "inventario/reporte_disponibilidad.html", context)

# --- CRUD PROVEEDORES ---

@login_required
@perfil_required(allow=("administrador", "Encargado"), readonly_for=("Bodeguero",))
def listar_proveedores(request):
    qs = models.Proveedor.objects.all().order_by("nombre_empresa")
    
    # Sort map simplified for default view
    sort_map = {
        "nombre": "nombre_empresa",
        "rut": "rut_empresa",
        "estado": "estado",
        "ciudad": "ciudad",
    }
    
    read_only = not (request.user.is_superuser or user_has_role(request.user, "Administrador"))

    return list_with_filters(
        request,
        qs,
        search_fields=["nombre_empresa", "rut_empresa", "email", "ciudad"],
        order_field="nombre_empresa", # Default order field
        session_prefix="proveedores",
        context_key="proveedores",
        full_template="inventario/listar_proveedores.html", # Plantilla de lista principal
        partial_template="inventario/partials/proveedores_results.html", # Tabla AJAX
        default_per_page=10,
        tie_break="id",
        extra_context={
            "titulo": "Listado de Proveedores",
            "read_only": read_only,
        },
    )

@login_required
@perfil_required(allow=("administrador", "Encargado"))
def crear_proveedor(request):
    form = ProveedorForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "✅ Proveedor creado correctamente.")
        return redirect("inventario:listar_proveedores")
    return render(request, "inventario/proveedor_form.html", {"form": form, "titulo": "Nuevo Proveedor"})

@login_required
@perfil_required(allow=("administrador", "Encargado"))
def editar_proveedor(request, pk):
    obj = get_object_or_404(models.Proveedor, pk=pk)
    form = ProveedorForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f"📝 Proveedor '{obj.nombre_empresa}' actualizado.")
        return redirect("inventario:listar_proveedores")
    return render(request, "inventario/proveedor_form.html", {"form": form, "titulo": f"Editar: {obj.nombre_empresa}"})

@login_required
@perfil_required(allow=("administrador", "Encargado"))
@transaction.atomic
def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(models.Proveedor, pk=pk)
    
    # Regla: Verificar si tiene lotes asociados
    tiene_lotes = models.InsumoLote.objects.filter(proveedor=proveedor).exists()

    if request.method == "POST":
        
        # --- LÓGICA DE VALIDACIÓN ---
        if tiene_lotes:
            error_msg = "No se puede desactivar/eliminar el proveedor: tiene lotes de insumos asociados."
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            messages.error(request, error_msg)
            return redirect("inventario:listar_proveedores")
            
        # --- LÓGICA DE ÉXITO ---
        proveedor.is_active = False 
        proveedor.save(update_fields=["is_active"])
        success_msg = f"Proveedor '{proveedor.nombre_empresa}' desactivado."
        
        # Respuesta AJAX: Devuelve éxito y mensaje
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({'success': True, 'message': success_msg}, status=200)

        # Respuesta de Formulario Normal: Redirecciona con mensaje
        messages.success(request, success_msg)
        return redirect("inventario:listar_proveedores")

    return render(
        request,
        "inventario/confirmar_eliminar.html",
        {"titulo": f"Desactivar Proveedor «{proveedor.nombre_empresa}»", "cancel_url": reverse("inventario:listar_proveedores"), "obj": proveedor}
    )
